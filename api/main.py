#!/usr/bin/env python3
# api/main.py  —  Vercel 서버리스 FastAPI
# 변경 사항 (backend/main.py 대비):
#   1. /tmp 디렉토리 사용 (Vercel 서버리스 파일쓰기)
#   2. 기업코드 XML /tmp 캐싱 (콜드스타트 외 재다운로드 생략)
#   3. ThreadPoolExecutor 병렬 DART 호출 (3회사×3년=9개 동시)
#   4. Claude CLI 미사용 환경 대응 (503 정상 반환)
import os, json, re, shutil, subprocess, threading
from dotenv import load_dotenv
import requests, zipfile, xml.etree.ElementTree as ET, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

# ── 파일 경로 (/tmp 우선, 없으면 현재 디렉터리) ───────────────────────────────
BASE_DIR    = "/tmp" if os.path.isdir("/tmp") else os.path.dirname(os.path.abspath(__file__))
CORP_CACHE  = os.path.join(BASE_DIR, "dart_corps.xml")
EXCEL_DIR   = BASE_DIR

# ── 환경변수 로드 (.env → 없으면 시스템 환경변수) ─────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '.env'))

# ── Constants ──────────────────────────────────────────────────────────────────
API_KEY  = os.getenv("DART_API_KEY", "")
CORP_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
FIN_URL  = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
YEARS    = ["2023", "2024", "2025"]

CO_COLS  = ["3182F6", "FF6B35", "7C3AED"]
LV_COLOR = {"good": "#05C072", "note": "#F59E0B", "warn": "#F97316", "danger": "#F04452"}
LV_XL    = {"good": "05C072",  "note": "F59E0B",  "warn": "F97316",  "danger": "F04452"}
LV_XL_BG = {"good": "E8FBF3",  "note": "FFFBEB",  "warn": "FFF4ED",  "danger": "FFF0F1"}
LV_ICON  = {"good": "✓", "note": "!", "warn": "△", "danger": "✕"}

OPINION_KEYS = [
    ("수익성", "영업이익률"),   ("수익성", "순이익률"),
    ("수익성", "매출총이익률"), ("효율성", "ROE"),
    ("효율성", "ROA"),          ("안정성", "부채비율"),
    ("안정성", "유동비율"),     ("안정성", "이자보상배율"),
    ("안정성", "이익잉여금 비율"),
    ("성장성", "매출 성장률"),  ("성장성", "영업이익 성장률"),
    ("성장성", "순이익 성장률"),("총평",   "종합 의견"),
]

INCOME_KEYS  = ["매출","매출원가","매출총이익","판관비","영업이익",
                "금융수익","금융비용","세전이익","법인세비용","당기순이익","EPS"]
BALANCE_KEYS = ["유동자산","비유동자산","자산총계",
                "유동부채","비유동부채","부채총계","이익잉여금","자본총계"]
RATIO_KEYS   = ["영업이익률","순이익률","ROE","ROA","부채비율","유동비율","이자보상배율"]
SECTIONS     = [
    ("손익계산서",  INCOME_KEYS,  False),
    ("재무상태표",  BALANCE_KEYS, False),
    ("수익성 지표", RATIO_KEYS,   True),
]
EPS_KEYS   = {"EPS"}
XRATE_KEYS = {"이자보상배율"}

# ── CODE_MAP (account_id → canonical) ─────────────────────────────────────────
CODE_MAP: dict = {
    "ifrs-full_Revenue":                                "매출",
    "ifrs-full_RevenueFromContractsWithCustomers":      "매출",
    "dart_Revenue":                                     "매출",
    "ifrs-full_CostOfSales":                            "매출원가",
    "dart_CostOfSales":                                 "매출원가",
    "ifrs-full_GrossProfit":                            "매출총이익",
    "ifrs-full_SellingGeneralAndAdministrativeExpense": "판관비",
    "dart_SellingGeneralAdministrativeExpenses":        "판관비",
    "ifrs-full_AdministrativeExpense":                  "판관비",
    "dart_OperatingIncomeLoss":                         "영업이익",
    "ifrs-full_ProfitLossFromOperatingActivities":      "영업이익",
    "ifrs-full_FinanceIncome":                          "금융수익",
    "ifrs-full_FinanceCosts":                           "금융비용",
    "ifrs-full_ProfitLossBeforeTax":                    "세전이익",
    "ifrs-full_IncomeTaxExpenseContinuingOperations":   "법인세비용",
    "ifrs-full_ProfitLoss":                             "당기순이익",
    "dart_ProfitLoss":                                  "당기순이익",
    "us-gaap_NetIncomeLoss":                            "당기순이익",
    "ifrs-full_BasicEarningsLossPerShare":              "EPS",
    "ifrs-full_BasicEarningsPerShare":                  "EPS",
    "dart_BasicEarningsLossPerShare":                   "EPS",
    "ifrs-full_CurrentAssets":                          "유동자산",
    "ifrs-full_NoncurrentAssets":                       "비유동자산",
    "ifrs-full_Assets":                                 "자산총계",
    "us-gaap_Assets":                                   "자산총계",
    "ifrs-full_CurrentLiabilities":                     "유동부채",
    "ifrs-full_NoncurrentLiabilities":                  "비유동부채",
    "ifrs-full_Liabilities":                            "부채총계",
    "ifrs-full_RetainedEarnings":                       "이익잉여금",
    "ifrs-full_Equity":                                 "자본총계",
    "us-gaap_StockholdersEquity":                       "자본총계",
    "ifrs-full_EquityAttributableToOwnersOfParent":     "자본총계",
}

# ── NAME_MAP (account_nm → canonical, fallback) ───────────────────────────────
NAME_MAP: dict = {
    "매출액": "매출", "수익(매출액)": "매출", "영업수익": "매출",
    "매출": "매출", "도급공사수익": "매출", "분양수익": "매출",
    "이자수익": "매출",
    "매출원가": "매출원가", "도급공사원가": "매출원가",
    "매출총이익": "매출총이익", "매출총이익(손실)": "매출총이익",
    "매출총손익": "매출총이익",
    "판매비와관리비": "판관비", "판매비및관리비": "판관비",
    "판매비와 관리비": "판관비", "영업비용": "판관비",
    "영업이익": "영업이익", "영업이익(손실)": "영업이익", "영업손익": "영업이익",
    "금융수익": "금융수익", "이자및배당금수익": "금융수익",
    "금융비용": "금융비용", "이자비용": "금융비용", "금융원가": "금융비용",
    "법인세비용차감전순이익": "세전이익",
    "법인세비용차감전순이익(손실)": "세전이익",
    "법인세차감전순이익": "세전이익",
    "법인세비용차감전순손익": "세전이익",
    "법인세비용": "법인세비용", "법인세수익": "법인세비용",
    "당기순이익": "당기순이익", "당기순이익(손실)": "당기순이익",
    "분기순이익": "당기순이익", "반기순이익": "당기순이익",
    "당기순손익": "당기순이익",
    "기본주당순이익": "EPS", "기본주당이익": "EPS",
    "기본주당순이익(손실)": "EPS", "보통주기본주당이익(손실)": "EPS",
    "유동자산": "유동자산",
    "비유동자산": "비유동자산", "고정자산": "비유동자산",
    "자산총계": "자산총계", "자산 합계": "자산총계",
    "유동부채": "유동부채",
    "비유동부채": "비유동부채", "고정부채": "비유동부채",
    "부채총계": "부채총계", "부채 합계": "부채총계",
    "이익잉여금": "이익잉여금", "이익잉여금(결손금)": "이익잉여금",
    "미처분이익잉여금": "이익잉여금",
    "자본총계": "자본총계", "자본 합계": "자본총계",
}

# ── Utilities ──────────────────────────────────────────────────────────────────
def _parse(s, as_float=False):
    try:
        s = str(s).replace(",", "").strip()
        if s and s not in ("-", ""):
            return float(s) if as_float else int(float(s))
    except Exception:
        pass
    return None


def fmt_num(n, key=""):
    if n is None: return "N/A"
    if key in EPS_KEYS: return f"{int(n):,}원"
    v = n / 1e8
    if abs(v) >= 10000: return f"{v/10000:,.1f}조"
    return f"{v:,.1f}억"


def fmt_ratio(v, key=""):
    if v is None: return "N/A"
    if key in XRATE_KEYS: return f"{v:.2f}x"
    return f"{v:.1f}%"


def fmt_val(v, key="", is_ratio=False):
    return fmt_ratio(v, key) if is_ratio else fmt_num(v, key)


def yoy_str(cur, prev):
    if cur is None or prev is None or prev == 0: return ""
    r = (cur - prev) / abs(prev) * 100
    if r > 0:   return f"▲{abs(r):.1f}%"
    elif r < 0: return f"▼{abs(r):.1f}%"
    return "─"


# ── DART Data Layer ────────────────────────────────────────────────────────────
class DART:
    def __init__(self):
        self.corps: dict = {}
        self.ready = False
        self._lock = threading.Lock()

    def load(self):
        # 1순위: /tmp 캐시
        if os.path.exists(CORP_CACHE):
            try:
                root = ET.parse(CORP_CACHE).getroot()
                self._parse_root(root)
                self.ready = True
                return
            except Exception:
                pass

        # 2순위: DART API 다운로드
        r = requests.get(CORP_URL, params={"crtfc_key": API_KEY}, timeout=30)
        r.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            with z.open(z.namelist()[0]) as f:
                xml_bytes = f.read()

        # /tmp에 캐시 저장 (실패해도 계속 진행)
        try:
            with open(CORP_CACHE, "wb") as f:
                f.write(xml_bytes)
        except Exception:
            pass

        self._parse_root(ET.fromstring(xml_bytes))
        self.ready = True

    def _parse_root(self, root):
        self.corps = {}
        for item in root.findall("list"):
            nm = item.findtext("corp_name", "").strip()
            if not nm: continue
            self.corps.setdefault(nm, []).append({
                "corp_code":  item.findtext("corp_code", "").strip(),
                "stock_code": item.findtext("stock_code", "").strip(),
            })

    def search(self, q: str) -> list:
        if not self.ready: return []
        ql = q.lower()
        return [
            {"corp_name": nm, **c}
            for nm, corps in self.corps.items()
            if ql in nm.lower()
            for c in corps
        ][:30]

    def financials(self, corp_code: str, year: str):
        for fs in ("CFS", "OFS"):
            try:
                r = requests.get(FIN_URL, params={
                    "crtfc_key": API_KEY, "corp_code": corp_code,
                    "bsns_year": year, "reprt_code": "11011", "fs_div": fs,
                }, timeout=30)
                d = r.json()
                if d.get("status") == "000" and d.get("list"):
                    return d["list"], fs
            except Exception:
                pass
        return None, None

    def extract(self, items: list) -> dict:
        if not items: return {}
        res: dict = {}
        for item in items:
            acct_id   = (item.get("account_id") or "").strip()
            acct_nm   = (item.get("account_nm") or "").strip()
            canonical = CODE_MAP.get(acct_id) or NAME_MAP.get(acct_nm)
            if not canonical: continue
            v = _parse(item.get("thstrm_amount", ""), as_float=(canonical in EPS_KEYS))
            if v is None: continue
            if canonical not in res or abs(v) > abs(res[canonical]):
                res[canonical] = v
        return res

    @staticmethod
    def calc_ratios(d: dict) -> dict:
        rev   = d.get("매출")      or 0
        op    = d.get("영업이익")
        net   = d.get("당기순이익")
        asset = d.get("자산총계")  or 0
        eq    = d.get("자본총계")  or 0
        dbt   = d.get("부채총계")  or 0
        ca    = d.get("유동자산") or 0
        cl    = d.get("유동부채") or 0
        fc    = d.get("금융비용") or 0
        gp    = d.get("매출총이익") or 0
        er    = d.get("이익잉여금") or 0
        return {
            "영업이익률":    op  / rev   * 100 if op  is not None and rev   else None,
            "순이익률":      net / rev   * 100 if net is not None and rev   else None,
            "ROE":          net / eq    * 100 if net is not None and eq    else None,
            "ROA":          net / asset * 100 if net is not None and asset else None,
            "부채비율":      dbt / eq    * 100 if eq                        else None,
            "유동비율":      ca  / cl    * 100 if cl                        else None,
            "이자보상배율":  op  / fc          if op  is not None and fc    else None,
            "매출총이익률":  gp  / rev   * 100 if gp  and rev               else None,
            "이익잉여금 비율": er / asset * 100 if er and asset              else None,
        }


# ── Claude CLI (서버리스에서는 미사용 → 503 정상 반환) ──────────────────────────
CLAUDE_PROMPT = """\
당신은 한국 주식 투자 전문 재무 분석가입니다.
아래 기업 재무 데이터를 분석하고, 투자자 관점의 의견을 JSON으로만 출력하세요.
JSON 이외의 텍스트(설명, 마크다운 등)는 절대 출력하지 마세요.

{fin_table}

다음 13개 항목을 분석하여 정확히 아래 JSON 형식으로 출력하세요:
{{
  "영업이익률":      {{"value":"최신값(예:11.9%)", "text":"투자관점 의견 25자이내", "level":"good|note|warn|danger"}},
  "순이익률":        {{"value":"...", "text":"...", "level":"..."}},
  "매출총이익률":    {{"value":"...", "text":"...", "level":"..."}},
  "ROE":            {{"value":"...", "text":"...", "level":"..."}},
  "ROA":            {{"value":"...", "text":"...", "level":"..."}},
  "부채비율":        {{"value":"...", "text":"...", "level":"..."}},
  "유동비율":        {{"value":"...", "text":"...", "level":"..."}},
  "이자보상배율":    {{"value":"...(예:5.2x)", "text":"...", "level":"..."}},
  "이익잉여금 비율": {{"value":"...", "text":"...", "level":"..."}},
  "매출 성장률":     {{"value":"▲X.X% 또는 ▼X.X%", "text":"...", "level":"..."}},
  "영업이익 성장률": {{"value":"...", "text":"...", "level":"..."}},
  "순이익 성장률":   {{"value":"...", "text":"...", "level":"..."}},
  "종합 의견":       {{"value":"A|B|C|D", "text":"종합 등급 한 줄 평가", "level":"..."}}
}}
level 기준 — good:양호/강점, note:보통/중립, warn:주의/약점, danger:위험/심각"""


def check_claude_cli() -> bool:
    return shutil.which("claude") is not None


def call_claude_cli(prompt: str, timeout: int = 120) -> str | None:
    for cmd in [["claude", "--print", "-"], ["claude", "-p"]]:
        try:
            result = subprocess.run(
                cmd, input=prompt, capture_output=True,
                text=True, timeout=timeout, encoding="utf-8",
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return None
        except Exception:
            continue
    return None


def _extract_json(text: str) -> dict | None:
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try: return json.loads(m.group(1))
        except Exception: pass
    start, end = text.find("{"), text.rfind("}")
    if start != -1 and end > start:
        try: return json.loads(text[start:end+1])
        except Exception: pass
    return None


def _normalize_opinions(raw: dict) -> dict:
    result = {}
    for cat, lbl in OPINION_KEYS:
        item = raw.get(lbl, {})
        if not isinstance(item, dict): item = {}
        result[lbl] = {
            "category": cat,
            "value":    str(item.get("value", "N/A")),
            "text":     str(item.get("text",  "분석 데이터 없음")),
            "level":    item.get("level", "note")
                        if item.get("level") in LV_COLOR else "note",
        }
    return result


def _build_fin_table(corp_name: str, yearly: dict) -> str:
    avail = [yr for yr in YEARS if yearly.get(yr)]
    if not avail:
        return f"회사명: {corp_name}\n데이터 없음"
    lines = [f"회사명: {corp_name}  |  분석 기준: {avail[-1]}년 결산\n"]

    def row(label, vals, suffix="억원", fmt_fn=None):
        parts = []
        for i, (yr, v) in enumerate(zip(avail, vals)):
            if v is None:
                parts.append(f"{yr}: N/A")
                continue
            s = fmt_fn(v) if fmt_fn else f"{v/1e8:,.1f}{suffix}"
            if i > 0 and vals[i-1] is not None and vals[i-1] != 0:
                r = (v - vals[i-1]) / abs(vals[i-1]) * 100
                arrow = f"▲{r:.1f}%" if r > 0 else f"▼{abs(r):.1f}%"
                parts.append(f"{yr}: {s} ({arrow})")
            else:
                parts.append(f"{yr}: {s}")
        return f"  {label:<18} " + "  ".join(parts)

    lines.append("[ 손익계산서 ] (단위: 억원)")
    for key in INCOME_KEYS:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        if key == "EPS":
            lines.append(row(key, vals, "원", lambda v: f"{int(v):,}원"))
        else:
            lines.append(row(key, vals))

    lines.append("\n[ 재무상태표 ] (단위: 억원)")
    for key in BALANCE_KEYS:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        lines.append(row(key, vals))

    lines.append("\n[ 수익성·안정성 지표 ]")
    for key in RATIO_KEYS + ["매출총이익률", "이익잉여금 비율"]:
        vals = [yearly.get(yr, {}).get(key) for yr in avail]
        fn = (lambda v: f"{v:.2f}x") if key in XRATE_KEYS else (lambda v: f"{v:.1f}%")
        lines.append(row(key, vals, "", fn))

    return "\n".join(lines)


def generate_opinions_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = CLAUDE_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    return _normalize_opinions(raw)




# ── Deal Intelligence Prompts ──────────────────────────────────────────────────
DEAL_SIGNAL_PROMPT = """당신은 VC/M&A/PE/IPO 전문 딜 어드바이저입니다.
아래 기업의 재무 데이터를 기반으로 현재 딜 투자 가치를 평가하고 JSON으로만 출력하세요.
JSON 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "signal":       "Strong Buy|Watch|Caution|Pass 중 하나",
  "signal_level": "good|note|warn|danger",
  "summary":      "핵심 한 줄 평가 (20자 이내)",
  "reasons":      ["근거 1 (25자 이내)", "근거 2", "근거 3"],
  "timing":       "지금이 딜 적기인 이유 또는 아닌 이유 (40자 이내)",
  "deal_types":   ["VC", "M&A", "PE", "IPO", "CFO" 중 해당 항목들]
}}
signal_level: Strong Buy→good, Watch→note, Caution→warn, Pass→danger"""

INDUSTRY_RADAR_PROMPT = """당신은 산업 분석 전문 투자 애널리스트입니다.
아래 기업의 재무 데이터를 기반으로 산업 포지셔닝을 분석하고 JSON으로만 출력하세요.
JSON 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "industry":         "추정 업종/산업명",
  "market_dynamics":  "시장 동학 한 줄 설명 (40자 이내)",
  "moat":             "경쟁 해자 한 줄 설명 (40자 이내)",
  "growth_vectors":   ["성장 동력 1 (25자 이내)", "성장 동력 2", "성장 동력 3"],
  "threats":          ["주요 위협 1 (25자 이내)", "주요 위협 2"],
  "positioning":      "Leader|Challenger|Niche|Vulnerable 중 하나",
  "positioning_text": "산업 내 포지셔닝 한 줄 평 (30자 이내)"
}}"""

DEAL_MEMO_PROMPT = """당신은 투자은행 시니어 어드바이저입니다.
아래 기업의 재무 데이터를 기반으로 {deal_type} 관점의 Deal Memo를 Markdown으로 작성하세요.
Markdown 이외 텍스트는 절대 출력하지 마세요.

{fin_table}

반드시 아래 구조를 따르세요:

# Deal Memo: {corp_name}

**딜 유형:** {deal_type} | **작성일:** {date} | **기밀 등급:** Confidential

---

## 1. Executive Summary

## 2. Why Now?

## 3. Company Overview

## 4. Industry View

## 5. Financial Checkpoints

## 6. Deal Angle

## 7. Key Due Diligence Questions

## 8. Risks

## 9. Next Actions"""


# ── Deal Intelligence: Mock Functions ─────────────────────────────────────────
def _deal_signal_mock(corp_name: str, yearly: dict) -> dict:
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default=None)
    if not latest_yr:
        return {"signal": "Pass", "signal_level": "danger",
                "summary": "재무 데이터 없음",
                "reasons": ["DART 데이터를 불러올 수 없습니다"],
                "timing": "데이터 확인 후 재시도 필요", "deal_types": []}
    d = yearly[latest_yr]
    op_margin = d.get("영업이익률")
    roe       = d.get("ROE")
    prev_yr   = str(int(latest_yr) - 1)
    rev = d.get("매출"); prev_rev = yearly.get(prev_yr, {}).get("매출")
    rev_growth = (rev - prev_rev) / abs(prev_rev) * 100 if rev and prev_rev and prev_rev != 0 else None

    score = 0; reasons = []
    if op_margin is not None:
        if op_margin >= 15:  score += 2; reasons.append(f"영업이익률 {op_margin:.1f}% — 수익성 우수")
        elif op_margin >= 5: score += 1; reasons.append(f"영업이익률 {op_margin:.1f}% — 안정적")
        else:                score -= 1; reasons.append(f"영업이익률 {op_margin:.1f}% — 수익성 개선 필요")
    if roe is not None:
        if roe >= 15:  score += 2; reasons.append(f"ROE {roe:.1f}% — 자본 효율 우수")
        elif roe >= 8: score += 1; reasons.append(f"ROE {roe:.1f}% — 보통 수준")
        else:          score -= 1; reasons.append(f"ROE {roe:.1f}% — 자본 효율 낮음")
    if rev_growth is not None:
        if rev_growth >= 20:  score += 2; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 고성장")
        elif rev_growth >= 5: score += 1; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 성장세")
        else:                 score -= 1; reasons.append(f"매출 성장률 {rev_growth:.1f}% — 성장 둔화")
    if not reasons:
        reasons.append(f"{latest_yr}년 재무 데이터 기반 평가")

    if score >= 4:   signal, level, timing, dtypes = "Strong Buy", "good",   "재무지표 전반 양호, 딜 검토 적기",       ["VC", "M&A", "PE", "IPO"]
    elif score >= 2: signal, level, timing, dtypes = "Watch",      "note",   "일부 지표 긍정적, 추가 실사 권고",       ["M&A", "PE"]
    elif score >= 0: signal, level, timing, dtypes = "Caution",    "warn",   "리스크 존재, 신중한 접근 필요",          ["PE"]
    else:            signal, level, timing, dtypes = "Pass",       "danger", "재무 구조 개선 후 재검토 권고",          []
    return {"signal": signal, "signal_level": level,
            "summary": f"{corp_name} — {signal}",
            "reasons": reasons[:3], "timing": timing, "deal_types": dtypes}


def _industry_radar_mock(corp_name: str, yearly: dict) -> dict:
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default=None)
    d = yearly.get(latest_yr, {}) if latest_yr else {}
    op_margin = d.get("영업이익률")
    if op_margin and op_margin >= 15:   pos, pos_txt = "Leader",     "높은 수익성으로 시장 우위 추정"
    elif op_margin and op_margin >= 5:  pos, pos_txt = "Challenger", "안정적 수익 기반의 도전자 포지션"
    else:                               pos, pos_txt = "Niche",      "수익성 개선 또는 틈새 전략 필요"
    return {
        "industry": "정보 없음 (DART 사업보고서 확인 필요)",
        "market_dynamics": "공개 재무 데이터만으로는 산업 동학 분석 한계",
        "moat": "재무비율 기반 경쟁 해자 추정 — 상세 실사 필요",
        "growth_vectors": ["매출 성장률 기반 성장 동력 추정", "사업보고서 검토 필요", "업종별 드라이버 확인 권장"],
        "threats": ["시장 포화 가능성", "외부 경쟁 압력 증가"],
        "positioning": pos,
        "positioning_text": pos_txt,
    }


def _deal_memo_mock(corp_name: str, deal_type: str, yearly: dict) -> str:
    from datetime import date as _date
    today = _date.today().strftime("%Y년 %m월 %d일")
    latest_yr = max((yr for yr in YEARS if yearly.get(yr)), default="N/A")
    d = yearly.get(latest_yr, {})

    def _f(v):
        if v is None: return "N/A"
        v2 = v / 1e8
        if abs(v2) >= 10000: return f"{v2/10000:.1f}조원"
        return f"{v2:.1f}억원"
    def _p(v): return f"{v:.1f}%" if v is not None else "N/A"

    return f"""# Deal Memo: {corp_name}

**딜 유형:** {deal_type} | **작성일:** {today} | **기밀 등급:** Confidential

---

## 1. Executive Summary

{corp_name}은(는) {latest_yr}년 기준 매출 {_f(d.get("매출"))}, 영업이익 {_f(d.get("영업이익"))}, 당기순이익 {_f(d.get("당기순이익"))}을 기록한 기업입니다. {deal_type} 관점의 본 메모는 DART 공시 재무 데이터 기반으로 작성된 예비 검토 자료이며, 본격적인 실사(DD) 이전 단계입니다.

> ⚠️ Claude CLI가 연결되지 않아 자동 생성된 Mock 메모입니다. Claude CLI 연동 시 AI가 전문적으로 작성합니다.

## 2. Why Now?

- {latest_yr}년 결산 데이터 공시 완료 — 초기 재무 검토 가능 시점
- {deal_type} 딜 구조 검토를 위한 기초 재무 데이터 확보
- 추가 실사(Due Diligence) 진행 전 예비 스크리닝 단계

## 3. Company Overview

- **기업명:** {corp_name}
- **분석 기준:** {latest_yr}년 연간 결산
- **총자산:** {_f(d.get("자산총계"))}
- **자본총계:** {_f(d.get("자본총계"))}
- **이익잉여금:** {_f(d.get("이익잉여금"))}

## 4. Industry View

공개 재무제표 데이터만으로는 세부 산업 분류가 어렵습니다. DART 사업보고서의 업종 코드 및 주요 사업 내용을 확인하여 정확한 산업 포지셔닝을 파악하세요.

## 5. Financial Checkpoints

| 지표 | {latest_yr}년 |
|------|--------------|
| 매출 | {_f(d.get("매출"))} |
| 영업이익 | {_f(d.get("영업이익"))} |
| 당기순이익 | {_f(d.get("당기순이익"))} |
| 영업이익률 | {_p(d.get("영업이익률"))} |
| ROE | {_p(d.get("ROE"))} |
| ROA | {_p(d.get("ROA"))} |
| 부채비율 | {_p(d.get("부채비율"))} |
| 유동비율 | {_p(d.get("유동비율"))} |

## 6. Deal Angle

{deal_type} 관점의 상세 딜 구조 분석은 Claude CLI 연동 후 AI 자동 생성을 권장합니다. 기초 재무 지표를 바탕으로 DCF/멀티플 기반 밸류에이션 작업을 우선 진행하세요.

## 7. Key Due Diligence Questions

1. 매출 구성 및 주요 고객사 현황 (매출 집중도 확인)
2. 미래 수주 잔고 및 파이프라인
3. 핵심 인력 구성 및 이탈 리스크
4. 기존 부채 구조, 만기 일정 및 리파이낸싱 계획
5. 법적 분쟁, 규제 리스크 및 지적재산권 현황
6. 관계사 거래 및 내부거래 현황
7. 환경/ESG 관련 잠재 비용 및 리스크

## 8. Risks

- **데이터 한계 (상):** 공시 재무제표 기반으로 관리 회계 정보 미포함
- **업종 불확실성 (중):** 세부 사업 내용 및 산업 동향 추가 확인 필요
- **시장 리스크 (중):** 금리, 환율 등 거시 경제 변수 영향 가능성
- **실사 리스크 (하):** 공시되지 않은 우발채무 또는 법적 이슈 존재 가능

## 9. Next Actions

1. DART 사업보고서 전문 다운로드 및 검토
2. 경영진 면담(Management Meeting) 일정 협의
3. 외부 법무/회계 자문사 선정 및 실사 범위 협의
4. 상세 재무 모델(DCF, 비교 멀티플) 작성
5. 초기 Term Sheet 초안 작성 및 내부 승인 요청

---
*본 메모는 DART 공시 재무 데이터 기반으로 자동 생성된 예비 자료입니다. 투자 결정 전 반드시 전문가 검토 및 실사를 진행하세요.*
"""


# ── Deal Intelligence: Claude Functions ───────────────────────────────────────
def generate_deal_signal_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = DEAL_SIGNAL_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    if raw.get("signal") not in {"Strong Buy", "Watch", "Caution", "Pass"}:
        raw["signal"] = "Watch"
    if raw.get("signal_level") not in {"good", "note", "warn", "danger"}:
        raw["signal_level"] = "note"
    if not isinstance(raw.get("reasons"), list):   raw["reasons"]    = []
    if not isinstance(raw.get("deal_types"), list): raw["deal_types"] = []
    return raw


def generate_industry_radar_claude(corp_name: str, yearly: dict) -> dict | None:
    if not check_claude_cli(): return None
    prompt   = INDUSTRY_RADAR_PROMPT.format(fin_table=_build_fin_table(corp_name, yearly))
    response = call_claude_cli(prompt)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    if raw.get("positioning") not in {"Leader", "Challenger", "Niche", "Vulnerable"}:
        raw["positioning"] = "Challenger"
    if not isinstance(raw.get("growth_vectors"), list): raw["growth_vectors"] = []
    if not isinstance(raw.get("threats"), list):        raw["threats"]        = []
    return raw


def generate_deal_memo_claude(corp_name: str, deal_type: str, yearly: dict) -> str | None:
    if not check_claude_cli(): return None
    from datetime import date as _date
    prompt = DEAL_MEMO_PROMPT.format(
        fin_table=_build_fin_table(corp_name, yearly),
        corp_name=corp_name, deal_type=deal_type,
        date=_date.today().strftime("%Y년 %m월 %d일"),
    )
    response = call_claude_cli(prompt, timeout=180)
    if not response: return None
    text = response.strip()
    if text.startswith("```markdown"): text = text[11:]
    elif text.startswith("```"):       text = text[3:]
    if text.endswith("```"):           text = text[:-3]
    return text.strip()




# ── 산업 인텔리전스: Naver News API ───────────────────────────────────────────
NAVER_CLIENT_ID     = os.getenv("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET", "")


def _extract_domain(url: str) -> str:
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc
        return host.replace("www.", "").split(".")[0] if host else ""
    except Exception:
        return ""


def search_naver_news(query: str, display: int = 15) -> list:
    """Naver News Search API. API 키 없으면 빈 리스트."""
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        return []
    try:
        r = requests.get(
            "https://openapi.naver.com/v1/search/news.json",
            headers={
                "X-Naver-Client-Id":     NAVER_CLIENT_ID,
                "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
            },
            params={"query": query, "display": display, "sort": "date"},
            timeout=10,
        )
        if r.status_code != 200:
            return []
        return [
            {
                "title":        re.sub(r'<[^>]+>', '', item.get("title", "")),
                "description":  re.sub(r'<[^>]+>', '', item.get("description", "")),
                "url":          item.get("originallink") or item.get("link", ""),
                "source":       _extract_domain(item.get("originallink") or item.get("link", "")),
                "published_at": item.get("pubDate", "")[:16],
            }
            for item in r.json().get("items", [])
        ]
    except Exception:
        return []


def _fmt_news_for_claude(news: list, text_input: str = "") -> str:
    parts = []
    for i, n in enumerate(news[:12], 1):
        parts.append(f"{i}. [{n.get('source','미상')}] {n['title']}")
        if n.get("description"):
            parts.append(f"   {n['description'][:150]}")
        if n.get("published_at"):
            parts.append(f"   ({n['published_at']})")
    if text_input and text_input.strip():
        parts.append("\n[사용자 입력 본문]")
        parts.append(text_input.strip()[:4000])
    return "\n".join(parts) if parts else "뉴스/본문 데이터 없음"


# ── 산업 인텔리전스: Claude 프롬프트 ──────────────────────────────────────────
COMPANY_TRENDS_PROMPT = """\
당신은 VC/M&A/PE 전문 산업 분석 어드바이저입니다.
아래 뉴스/자료와 재무 데이터를 기반으로 {corp_name}이 속한 산업 동향과 딜 인사이트를 분석하세요.
JSON으로만 출력하세요. JSON 이외 텍스트는 절대 출력하지 마세요.

[뉴스/자료]
{news_text}

[재무 데이터]
{fin_table}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "core_signals":           ["핵심 산업 시그널 1 (25자)", "시그널 2", "시그널 3"],
  "recent_issues":          [{{"title":"이슈명(20자)","summary":"30자 요약"}}, ...],
  "impact_on_corp":         "{corp_name}에 미치는 영향 (60자)",
  "deal_angle":             "딜 관점 코멘트 VC/M&A/PE 중 해당 관점 (50자)",
  "financial_impact":       "재무지표 변화 예상 (40자)",
  "accounting_checkpoints": ["회계·공시 체크포인트 1 (25자)", "2", "3"],
  "dd_questions":           ["실사 필수 질문 1 (30자)", "2", "3"],
  "risks":                  [{{"factor":"리스크명(20자)","level":"상|중|하"}}]
}}"""

INDUSTRY_ANALYSIS_PROMPT = """\
당신은 투자은행 시니어 산업 리서치 애널리스트입니다.
아래 뉴스/자료를 기반으로 {industry} 산업을 종합 분석하세요.
JSON으로만 출력하세요. JSON 이외 텍스트는 절대 출력하지 마세요.

[뉴스/자료]
{news_text}

아래 JSON 형식으로 정확히 출력하세요:
{{
  "sector_overview":    "섹터 개요 한 문장 (70자)",
  "recent_trends":      ["최근 트렌드 1 (30자)", "2", "3"],
  "growth_drivers":     ["성장 동력 1 (30자)", "2", "3"],
  "risks":              ["주요 리스크 1 (30자)", "2", "3"],
  "major_players":      ["주요 플레이어/세그먼트 1", "2", "3"],
  "deal_opportunities": ["딜 기회 1 (40자)", "딜 기회 2"],
  "valuation_points":   ["밸류에이션 포인트 1 (35자)", "포인트 2"],
  "accounting_points":  ["회계 포인트 1 (35자)", "포인트 2"],
  "dd_questions":       ["실사 질문 1 (35자)", "질문 2", "질문 3"]
}}"""


# ── 산업 인텔리전스: Mock ──────────────────────────────────────────────────────
def _company_trends_mock(corp_name: str, industry: str, has_data: bool) -> dict:
    hint = "" if has_data else " (뉴스 API 키 또는 본문 입력 필요)"
    return {
        "core_signals": [
            f"{industry or corp_name + ' 관련 산업'} 동향 분석 준비 중{hint}",
            "NAVER_CLIENT_ID/SECRET .env 설정 시 실제 뉴스 자동 수집",
            "Claude CLI 연동 시 AI 딜 분석 자동 생성",
        ],
        "recent_issues": [{"title": "데이터 없음", "summary": "뉴스 API 키 또는 본문 직접 입력 필요"}],
        "impact_on_corp": f"{corp_name}에 대한 분석은 뉴스 데이터 수집 후 제공됩니다",
        "deal_angle": "뉴스 수집 후 VC/M&A/PE 관점 딜 각도 분석 제공",
        "financial_impact": "데이터 입력 후 재무 영향 자동 분석",
        "accounting_checkpoints": ["수익 인식 정책 검토", "주요 자산 손상 여부 확인", "우발채무 공시 검토"],
        "dd_questions": ["핵심 사업 모델 및 경쟁 우위는?", "최근 3년 매출 구성 변화 원인은?", "주요 고객사 집중도 및 계약 현황은?"],
        "risks": [{"factor": "데이터 부족", "level": "중"}, {"factor": "시장 불확실성", "level": "중"}],
    }


def _industry_analysis_mock(industry: str, has_data: bool) -> dict:
    hint = "" if has_data else " (뉴스 API 키 또는 본문 입력 필요)"
    return {
        "sector_overview": f"{industry} 산업 분석을 위해 뉴스 데이터 입력이 필요합니다{hint}",
        "recent_trends":      [f"{industry} 트렌드 수집 필요", "NAVER API 키 설정 권장", "또는 기사 본문 직접 붙여넣기"],
        "growth_drivers":     ["데이터 입력 후 자동 추출", "Naver API 또는 수동 입력", "Claude CLI 연동 필요"],
        "risks":              ["데이터 부족으로 분석 불가", "뉴스 API 키 설정 필요", "Claude CLI 미연결"],
        "major_players":      ["데이터 입력 후 분석 가능"],
        "deal_opportunities": ["뉴스 분석 후 제공"],
        "valuation_points":   ["데이터 입력 후 분석"],
        "accounting_points":  ["데이터 입력 후 분석"],
        "dd_questions":       ["산업 내 경쟁 구도 변화는?", "주요 성장 드라이버 지속 가능성은?", "규제 환경 변화 리스크는?"],
    }


# ── 산업 인텔리전스: Claude 생성 함수 ─────────────────────────────────────────
def generate_company_trends_claude(corp_name: str, news: list, text_input: str, fin_table: str) -> dict | None:
    if not check_claude_cli(): return None
    news_text = _fmt_news_for_claude(news, text_input)
    if news_text == "뉴스/본문 데이터 없음": return None
    prompt = COMPANY_TRENDS_PROMPT.format(
        corp_name=corp_name, news_text=news_text,
        fin_table=fin_table or "재무 데이터 없음",
    )
    response = call_claude_cli(prompt, timeout=120)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    for k in ["core_signals", "recent_issues", "accounting_checkpoints", "dd_questions", "risks"]:
        if not isinstance(raw.get(k), list): raw[k] = []
    return raw


def generate_industry_analysis_claude(industry: str, news: list, text_input: str) -> dict | None:
    if not check_claude_cli(): return None
    news_text = _fmt_news_for_claude(news, text_input)
    if news_text == "뉴스/본문 데이터 없음": return None
    prompt = INDUSTRY_ANALYSIS_PROMPT.format(industry=industry, news_text=news_text)
    response = call_claude_cli(prompt, timeout=120)
    if not response: return None
    raw = _extract_json(response)
    if not raw: return None
    for k in ["recent_trends","growth_drivers","risks","major_players","deal_opportunities","valuation_points","accounting_points","dd_questions"]:
        if not isinstance(raw.get(k), list): raw[k] = []
    return raw


# ── Excel Generation ───────────────────────────────────────────────────────────
def _xl_styles():
    def mf(bold=False, color="191F28", size=11, italic=False):
        return Font(name="맑은 고딕", bold=bold, color=color, size=size, italic=italic)
    def fl(hex_color):
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))
    BD   = Border(
        left=Side(style="thin", color="E5E8EB"),
        right=Side(style="thin", color="E5E8EB"),
        top=Side(style="thin", color="E5E8EB"),
        bottom=Side(style="thin", color="E5E8EB"),
    )
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    return mf, fl, BD, CTR, LEFT


def _write_company_sheet(ws, ci, corp_name, yearly_ci):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    co_hex = CO_COLS[ci]
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    ws.merge_cells("A1:G2")
    cell(1, 1, f"{corp_name}  재무분석", mf(True,"FFFFFF",14), co_hex, CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells("A3:G3")
    cell(3, 1, f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    for sec_title, keys, is_ratio in SECTIONS:
        ws.merge_cells(f"A{r}:G{r}")
        cell(r, 1, sec_title, mf(True,"FFFFFF",12), co_hex, CTR)
        ws.row_dimensions[r].height = 26; r += 1
        for ci2, h in enumerate(["항목","2023","2024","YoY","2025","YoY",""], 1):
            cell(r, ci2, h, mf(True,"FFFFFF",10), "6B7684", CTR)
        ws.row_dimensions[r].height = 20; r += 1
        for ki, key in enumerate(keys):
            vals = [yearly_ci.get(yr, {}).get(key)
                    if yearly_ci.get(yr) else None for yr in YEARS]
            bg = "F9FAFB" if ki % 2 == 0 else "FFFFFF"
            ws.row_dimensions[r].height = 22
            cell(r, 1, key, mf(color="6B7684"), bg, LEFT)
            cell(r, 2, fmt_val(vals[0], key, is_ratio), mf(True), bg, CTR)
            cell(r, 7, "", mf(), bg, CTR)
            for col_off, idx in [(3, 1), (5, 2)]:
                v   = vals[idx]
                txt = yoy_str(v, vals[idx-1])
                cell(r, col_off, fmt_val(v, key, is_ratio), mf(True), bg, CTR)
                yf  = (mf(color="05C072",size=10) if txt.startswith("▲")
                       else mf(color="F04452",size=10) if txt.startswith("▼")
                       else mf(color="B0B8C1",size=10))
                cell(r, col_off+1, txt or "-", yf, bg, CTR)
            r += 1
        r += 1
    for col_l, w in [("A",14),("B",14),("C",14),("D",11),("E",14),("F",11),("G",6)]:
        ws.column_dimensions[col_l].width = w


def _write_comparison_sheet(ws, cos, yearly):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    active   = [(ci, c["corp_name"]) for ci, c in enumerate(cos) if c]
    n        = len(active)
    last_col = 1 + n * 3 + 1
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    ws.merge_cells(f"A1:{chr(64+last_col)}2")
    cell(1, 1, "3사 재무 비교 분석", mf(True,"FFFFFF",14), "191F28", CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A3:{chr(64+last_col)}3")
    cell(3, 1, f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    for sec_title, keys, is_ratio in SECTIONS:
        ws.merge_cells(f"A{r}:{chr(64+last_col)}{r}")
        cell(r, 1, sec_title, mf(True,"FFFFFF",12), "4B9FFF", CTR)
        ws.row_dimensions[r].height = 26; r += 1
        cell(r, 1, "항목", mf(True,"FFFFFF",10), "6B7684", LEFT)
        for li, (ci, nm) in enumerate(active):
            base_c = 2 + li * 3
            ws.merge_cells(f"{chr(64+base_c)}{r}:{chr(64+base_c+2)}{r}")
            cell(r, base_c, nm, mf(True,"FFFFFF",11), CO_COLS[ci], CTR)
        ws.row_dimensions[r].height = 26; r += 1
        cell(r, 1, "", mf(True,"FFFFFF",9), "6B7684", CTR)
        for li, (ci, _) in enumerate(active):
            for yi, yr in enumerate(YEARS):
                cell(r, 2+li*3+yi, yr, mf(True,"FFFFFF",9), "6B7684", CTR)
        ws.row_dimensions[r].height = 18; r += 1
        for ki, key in enumerate(keys):
            bg = "F9FAFB" if ki % 2 == 0 else "FFFFFF"
            ws.row_dimensions[r].height = 22
            cell(r, 1, key, mf(color="6B7684"), bg, LEFT)
            for li, (ci, _) in enumerate(active):
                for yi, yr in enumerate(YEARS):
                    v = yearly[ci].get(yr, {}).get(key) if yearly[ci].get(yr) else None
                    cell(r, 2+li*3+yi, fmt_val(v, key, is_ratio), mf(True), bg, CTR)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 14
    for li in range(n):
        for yi in range(3):
            ws.column_dimensions[chr(66+li*3+yi)].width = 13


def _write_opinion_sheet(ws, cos, opinions):
    mf, fl, BD, CTR, LEFT = _xl_styles()
    active = [(ci, c["corp_name"]) for ci, c in enumerate(cos) if c]
    n = len(active)
    def cell(r, c, v="", font=None, bg=None, align=CTR):
        cc = ws.cell(r, c, v)
        if font: cc.font = font
        if bg:   cc.fill = fl(bg)
        cc.alignment = align; cc.border = BD
    last_c = 2 + n
    ws.merge_cells(f"A1:{chr(64+last_c)}2")
    cell(1, 1, "재무 분석 의견 — Claude AI 자동 평가",
         mf(True,"FFFFFF",14), "191F28", CTR)
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A3:{chr(64+last_c)}3")
    cell(3, 1, f"생성일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  Claude CLI",
         mf(color="6B7684",size=10,italic=True), "FFFFFF", LEFT)
    ws.row_dimensions[3].height = 16
    r = 5
    cell(r, 1, "분류",    mf(True,"FFFFFF",10), "6B7684", CTR)
    cell(r, 2, "평가 항목", mf(True,"FFFFFF",10), "6B7684", CTR)
    for li, (ci, nm) in enumerate(active):
        cell(r, 3+li, nm, mf(True,"FFFFFF",11), CO_COLS[ci], CTR)
    ws.row_dimensions[r].height = 22; r += 1
    prev_cat = None
    for cat, lbl in OPINION_KEYS:
        bg_lbl = "EFF6FF" if cat != prev_cat else "FFFFFF"
        prev_cat = cat
        ws.row_dimensions[r].height = 54
        cell(r, 1, cat, mf(True,"6B7684",10), bg_lbl, CTR)
        cell(r, 2, lbl, mf(bold=True,size=11), "FFFFFF", LEFT)
        for li, (ci, _) in enumerate(active):
            op   = opinions[ci] or {}
            item = op.get(lbl, {"value":"N/A","text":"데이터 없음","level":"note"})
            lv   = item.get("level","note")
            combined = f"{LV_ICON.get(lv,'·')} {item.get('value','N/A')}\n{item.get('text','')}"
            cc   = ws.cell(r, 3+li, combined)
            cc.font      = Font(name="맑은 고딕", size=10, color=LV_XL.get(lv,"B0B8C1"),
                                bold=(lv=="good"))
            cc.fill      = PatternFill("solid", fgColor=LV_XL_BG.get(lv,"FFFFFF"))
            cc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cc.border    = BD
        r += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for li in range(n):
        ws.column_dimensions[chr(67+li)].width = 44


def generate_excel(cos: list, yearly: list, opinions: list | None = None) -> str:
    if opinions is None:
        opinions = [None, None, None]
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    names    = [c["corp_name"] for c in cos if c]
    stem     = "_".join(names[:2]) + ("_외" if len(names) > 2 else "")
    filename = f"{stem}_재무비교_{ts}.xlsx"
    path     = os.path.join(EXCEL_DIR, filename)
    wb       = openpyxl.Workbook(); wb.remove(wb.active)
    for ci, corp in enumerate(cos):
        if not corp: continue
        ws = wb.create_sheet(title=corp["corp_name"][:28])
        _write_company_sheet(ws, ci, corp["corp_name"], yearly[ci])
    ws_cmp = wb.create_sheet(title="전체 비교")
    _write_comparison_sheet(ws_cmp, cos, yearly)
    if any(opinions):
        ws_op = wb.create_sheet(title="재무분석 의견")
        _write_opinion_sheet(ws_op, cos, opinions)
    for target in reversed(["재무분석 의견", "전체 비교"]):
        if target in wb.sheetnames:
            wb.move_sheet(target, offset=-len(wb.sheetnames))
    wb.save(path)
    return filename


# ── FastAPI App ────────────────────────────────────────────────────────────────
app = FastAPI(title="DART 재무분석기 API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_dart       = DART()
_dart_lock  = threading.Lock()
_dart_ready = False


def _ensure_dart():
    if not API_KEY:
        raise HTTPException(status_code=500, detail="DART_API_KEY가 설정되지 않았습니다. .env 파일을 확인하세요.")
    global _dart_ready
    if not _dart_ready:
        with _dart_lock:
            if not _dart_ready:
                _dart.load()
                _dart_ready = True


# ── Pydantic Models ────────────────────────────────────────────────────────────
class SearchReq(BaseModel):
    query: str

class CorpInfo(BaseModel):
    corp_name:  str
    corp_code:  str
    stock_code: str = ""

class AnalyzeReq(BaseModel):
    corps: list[Optional[CorpInfo]]

class OpinionReq(BaseModel):
    corp_name: str
    yearly:    dict


# ── Parallel DART Fetcher ──────────────────────────────────────────────────────
def _fetch_parallel(cos: list) -> list:
    """3회사 × 3년 = 최대 9개 DART API 동시 호출. 순차 대비 ~3배 속도."""
    yearly = [{}, {}, {}]
    tasks  = [
        (ci, corp["corp_code"], yr)
        for ci, corp in enumerate(cos) if corp
        for yr in YEARS
    ]
    if not tasks:
        return yearly

    with ThreadPoolExecutor(max_workers=min(9, len(tasks))) as ex:
        futures = {
            ex.submit(_dart.financials, corp_code, yr): (ci, yr)
            for ci, corp_code, yr in tasks
        }
        for future in as_completed(futures):
            ci, yr = futures[future]
            try:
                items, fs = future.result()
            except Exception:
                continue
            if items:
                fin = _dart.extract(items)
                rat = DART.calc_ratios(fin)
                yearly[ci][yr] = {**fin, **rat, "_fs": fs}

    return yearly


# ── Endpoints ──────────────────────────────────────────────────────────────────
_pool = ThreadPoolExecutor(max_workers=4)


@app.post("/api/search")
async def search_corp(body: SearchReq):
    import asyncio
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(_pool, _ensure_dart)
    return {"results": _dart.search(body.query)}


@app.post("/api/analyze-multi")
async def analyze_multi(body: AnalyzeReq):
    import asyncio
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(_pool, _ensure_dart)

    def _run():
        cos    = [c.model_dump() if c else None for c in body.corps]
        yearly = _fetch_parallel(cos)          # ← 병렬 호출
        filename = generate_excel(cos, yearly)
        return {"yearly": yearly, "filename": filename}

    return await loop.run_in_executor(_pool, _run)


@app.post("/api/opinion")
async def get_opinion(body: OpinionReq):
    if not check_claude_cli():
        raise HTTPException(status_code=503, detail="Claude CLI를 사용할 수 없는 환경입니다")
    import asyncio
    loop   = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        _pool, lambda: generate_opinions_claude(body.corp_name, body.yearly)
    )
    if result is None:
        raise HTTPException(status_code=504, detail="Claude CLI 응답 없음")
    return result


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    safe = os.path.basename(filename)                    # 경로 순회 방지
    path = os.path.join(EXCEL_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")
    return FileResponse(
        path, filename=safe,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




# ── Deal Intelligence Models ───────────────────────────────────────────────────
class DealSignalReq(BaseModel):
    corp_name: str
    yearly:    dict

class IndustryRadarReq(BaseModel):
    corp_name: str
    yearly:    dict

class DealMemoReq(BaseModel):
    corp_name: str
    deal_type: str = "M&A"
    yearly:    dict


# ── Deal Intelligence Endpoints ────────────────────────────────────────────────
@app.post("/api/deal-signal")
async def deal_signal(body: DealSignalReq):
    import asyncio
    loop = asyncio.get_event_loop()
    def _run():
        result = generate_deal_signal_claude(body.corp_name, body.yearly)
        return result or _deal_signal_mock(body.corp_name, body.yearly)
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/industry-radar")
async def industry_radar(body: IndustryRadarReq):
    import asyncio
    loop = asyncio.get_event_loop()
    def _run():
        result = generate_industry_radar_claude(body.corp_name, body.yearly)
        return result or _industry_radar_mock(body.corp_name, body.yearly)
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/deal-memo")
async def deal_memo_endpoint(body: DealMemoReq):
    import asyncio
    loop = asyncio.get_event_loop()
    def _run():
        memo_text = generate_deal_memo_claude(body.corp_name, body.deal_type, body.yearly)
        if not memo_text:
            memo_text = _deal_memo_mock(body.corp_name, body.deal_type, body.yearly)
        memo_dir = os.path.join(BASE_DIR, "memos")
        os.makedirs(memo_dir, exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_nm  = re.sub(r'[^\w가-힣-]', '_', body.corp_name)
        filename = f"{safe_nm}_{body.deal_type}_{ts}.md"
        with open(os.path.join(memo_dir, filename), "w", encoding="utf-8") as fh:
            fh.write(memo_text)
        return {"memo": memo_text, "filename": filename}
    return await loop.run_in_executor(_pool, _run)


@app.get("/api/download-memo/{filename}")
async def download_memo_file(filename: str):
    safe = os.path.basename(filename)
    path = os.path.join(BASE_DIR, "memos", safe)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="메모 파일을 찾을 수 없습니다")
    return FileResponse(path, filename=safe, media_type="text/markdown; charset=utf-8")




# ── 산업 인텔리전스: Pydantic Models ──────────────────────────────────────────
class CompanyTrendsReq(BaseModel):
    corp_name:  str
    industry:   str  = ""
    text_input: str  = ""
    yearly:     dict = {}

class IndustryAnalysisReq(BaseModel):
    industry:   str
    text_input: str = ""


# ── 산업 인텔리전스: Endpoints ─────────────────────────────────────────────────
@app.post("/api/company-industry-trends")
async def company_industry_trends(body: CompanyTrendsReq):
    loop = asyncio.get_event_loop()
    def _run():
        query    = f"{body.corp_name} {body.industry}".strip()
        news     = search_naver_news(query, display=15)
        has_data = bool(news) or bool(body.text_input.strip())
        fin_table = _build_fin_table(body.corp_name, body.yearly) if body.yearly else ""
        result   = generate_company_trends_claude(body.corp_name, news, body.text_input, fin_table)
        if not result:
            result = _company_trends_mock(body.corp_name, body.industry, has_data)
        return {**result, "news": news, "naver_available": bool(NAVER_CLIENT_ID)}
    return await loop.run_in_executor(_pool, _run)


@app.post("/api/industry-analysis")
async def industry_analysis(body: IndustryAnalysisReq):
    loop = asyncio.get_event_loop()
    def _run():
        news     = search_naver_news(body.industry, display=15)
        has_data = bool(news) or bool(body.text_input.strip())
        result   = generate_industry_analysis_claude(body.industry, news, body.text_input)
        if not result:
            result = _industry_analysis_mock(body.industry, has_data)
        return {**result, "sources": news, "naver_available": bool(NAVER_CLIENT_ID)}
    return await loop.run_in_executor(_pool, _run)


@app.get("/api/health")
async def health():
    return {
        "status": "ok",
        "dart_ready": _dart_ready,
        "claude_cli": check_claude_cli(),
        "base_dir": BASE_DIR,
    }
