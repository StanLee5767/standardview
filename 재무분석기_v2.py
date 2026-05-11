#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DART 기업 재무분석기 v2 — 3사 동시 비교 + Claude CLI 의견

GUI 모드 : python 재무분석기_v2.py
CLI 모드 : python 재무분석기_v2.py data.json
"""

import customtkinter as ctk
from dotenv import load_dotenv
import tkinter as tk
import requests, zipfile, xml.etree.ElementTree as ET
import io, os, sys, json, re, shutil, subprocess, tempfile, threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ── Constants ──────────────────────────────────────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
API_KEY  = os.getenv("DART_API_KEY", "")
CORP_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
FIN_URL  = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
YEARS    = ["2023", "2024", "2025"]

C = {
    "bg":   "#F2F4F6", "card": "#FFFFFF",
    "blue": "#3182F6", "bdark":"#1B64DA",
    "grn":  "#05C072", "red":  "#F04452",
    "blk":  "#191F28", "g1":   "#6B7684",
    "g2":   "#B0B8C1", "g3":   "#E5E8EB", "g4": "#F9FAFB",
}
CO_COLS  = ["#3182F6", "#FF6B35", "#7C3AED"]
CO_DARKS = ["#1B64DA", "#E55320", "#5B21B6"]
FONT     = "Apple SD Gothic Neo"

LV_COLOR = {"good": "#05C072", "note": "#F59E0B", "warn": "#F97316", "danger": "#F04452"}
LV_BG    = {"good": "#E8FBF3", "note": "#FFFBEB", "warn": "#FFF4ED", "danger": "#FFF0F1"}
LV_ICON  = {"good": "✓", "note": "!", "warn": "△", "danger": "✕"}
LV_XL    = {"good": "05C072",  "note": "F59E0B",  "warn": "F97316",  "danger": "F04452"}
LV_XL_BG = {"good": "E8FBF3",  "note": "FFFBEB",  "warn": "FFF4ED",  "danger": "FFF0F1"}

OPINION_KEYS = [
    ("수익성", "영업이익률"),
    ("수익성", "순이익률"),
    ("수익성", "매출총이익률"),
    ("효율성", "ROE"),
    ("효율성", "ROA"),
    ("안정성", "부채비율"),
    ("안정성", "유동비율"),
    ("안정성", "이자보상배율"),
    ("안정성", "이익잉여금 비율"),
    ("성장성", "매출 성장률"),
    ("성장성", "영업이익 성장률"),
    ("성장성", "순이익 성장률"),
    ("총평",   "종합 의견"),
]

# ── account_id → 표준항목 (1순위) ─────────────────────────────────────────────
CODE_MAP: dict = {
    # 손익계산서
    "ifrs-full_Revenue":                                "매출",
    "ifrs-full_RevenueFromContractsWithCustomers":      "매출",
    "dart_Revenue":                                     "매출",
    "ifrs-full_CostOfSales":                            "매출원가",
    "dart_CostOfSales":                                 "매출원가",
    "ifrs-full_GrossProfit":                            "매출총이익",
    "ifrs-full_SellingGeneralAndAdministrativeExpense": "판관비",
    "dart_SellingGeneralAdministrativeExpenses":        "판관비",
    "ifrs-full_AdministrativeExpense":                  "판관비",
    "dart_OperatingIncomeLoss":                         "영업이익",
    "ifrs-full_ProfitLossFromOperatingActivities":      "영업이익",
    "ifrs-full_FinanceIncome":                          "금융수익",
    "ifrs-full_FinanceCosts":                           "금융비용",
    "ifrs-full_ProfitLossBeforeTax":                    "세전이익",
    "ifrs-full_IncomeTaxExpenseContinuingOperations":   "법인세비용",
    "ifrs-full_ProfitLoss":                             "당기순이익",
    "dart_ProfitLoss":                                  "당기순이익",
    "us-gaap_NetIncomeLoss":                            "당기순이익",
    "ifrs-full_BasicEarningsLossPerShare":              "EPS",
    "ifrs-full_BasicEarningsPerShare":                  "EPS",
    "dart_BasicEarningsLossPerShare":                   "EPS",
    # 재무상태표
    "ifrs-full_CurrentAssets":                          "유동자산",
    "ifrs-full_NoncurrentAssets":                       "비유동자산",
    "ifrs-full_Assets":                                 "자산총계",
    "us-gaap_Assets":                                   "자산총계",
    "ifrs-full_CurrentLiabilities":                     "유동부채",
    "ifrs-full_NoncurrentLiabilities":                  "비유동부채",
    "ifrs-full_Liabilities":                            "부채총계",
    "ifrs-full_RetainedEarnings":                       "이익잉여금",
    "ifrs-full_Equity":                                 "자본총계",
    "us-gaap_StockholdersEquity":                       "자본총계",
    "ifrs-full_EquityAttributableToOwnersOfParent":     "자본총계",
}

# ── account_nm → 표준항목 (2순위 폴백) ────────────────────────────────────────
NAME_MAP: dict = {
    # 매출
    "매출액": "매출", "수익(매출액)": "매출", "영업수익": "매출",
    "매출": "매출", "도급공사수익": "매출", "분양수익": "매출",
    "이자수익": "매출",   # 금융업 폴백
    # 매출원가
    "매출원가": "매출원가", "도급공사원가": "매출원가",
    # 매출총이익
    "매출총이익": "매출총이익", "매출총이익(손실)": "매출총이익",
    "매출총손익": "매출총이익",
    # 판관비
    "판매비와관리비": "판관비", "판매비및관리비": "판관비",
    "판매비와 관리비": "판관비", "영업비용": "판관비",
    # 영업이익
    "영업이익": "영업이익", "영업이익(손실)": "영업이익",
    "영업손익": "영업이익",
    # 금융수익·비용
    "금융수익": "금융수익", "이자및배당금수익": "금융수익",
    "금융비용": "금융비용", "이자비용": "금융비용", "금융원가": "금융비용",
    # 세전이익
    "법인세비용차감전순이익": "세전이익",
    "법인세비용차감전순이익(손실)": "세전이익",
    "법인세차감전순이익": "세전이익",
    "법인세비용차감전순손익": "세전이익",
    # 법인세
    "법인세비용": "법인세비용", "법인세수익": "법인세비용",
    # 당기순이익
    "당기순이익": "당기순이익", "당기순이익(손실)": "당기순이익",
    "분기순이익": "당기순이익", "반기순이익": "당기순이익",
    "당기순손익": "당기순이익",
    # EPS
    "기본주당순이익": "EPS", "기본주당이익": "EPS",
    "기본주당순이익(손실)": "EPS", "보통주기본주당이익(손실)": "EPS",
    # 자산
    "유동자산": "유동자산",
    "비유동자산": "비유동자산", "고정자산": "비유동자산",
    "자산총계": "자산총계", "자산 합계": "자산총계",
    # 부채
    "유동부채": "유동부채",
    "비유동부채": "비유동부채", "고정부채": "비유동부채",
    "부채총계": "부채총계", "부채 합계": "부채총계",
    # 자본
    "이익잉여금": "이익잉여금", "이익잉여금(결손금)": "이익잉여금",
    "미처분이익잉여금": "이익잉여금",
    "자본총계": "자본총계", "자본 합계": "자본총계",
}

INCOME_KEYS  = ["매출","매출원가","매출총이익","판관비","영업이익",
                "금융수익","금융비용","세전이익","법인세비용","당기순이익","EPS"]
BALANCE_KEYS = ["유동자산","비유동자산","자산총계",
                "유동부채","비유동부채","부채총계","이익잉여금","자본총계"]
RATIO_KEYS   = ["영업이익률","순이익률","ROE","ROA",
                "부채비율","유동비율","이자보상배율"]
SECTIONS     = [
    ("손익계산서",  INCOME_KEYS,  False),
    ("재무상태표",  BALANCE_KEYS, False),
    ("수익성 지표", RATIO_KEYS,   True),
]
EPS_KEYS   = {"EPS"}
XRATE_KEYS = {"이자보상배율"}

# ── Utilities ──────────────────────────────────────────────────────────────────
def _parse(s, as_float=False):
    try:
        s = str(s).replace(",", "").strip()
        if s and s not in ("-", ""):
            return float(s) if as_float else int(float(s))
    except Exception:
        pass
    return None


def fmt_num(n, key=""):
    if n is None: return "N/A"
    if key in EPS_KEYS: return f"{int(n):,}원"
    v = n / 1e8
    if abs(v) >= 10000: return f"{v/10000:,.1f}조"
    return f"{v:,.1f}억"


def fmt_ratio(v, key=""):
    if v is None: return "N/A"
    if key in XRATE_KEYS: return f"{v:.2f}x"
    return f"{v:.1f}%"


def fmt_val(v, key="", is_ratio=False):
    return fmt_ratio(v, key) if is_ratio else fmt_num(v, key)


def yoy(cur, prev):
    if cur is None or prev is None or prev == 0: return "", None
    r = (cur - prev) / abs(prev) * 100
    if r > 0:   return f"▲{abs(r):.1f}%", C["grn"]
    elif r < 0: return f"▼{abs(r):.1f}%", C["red"]
    return "─", C["g1"]


# ── Claude CLI Integration ─────────────────────────────────────────────────────
CLAUDE_INSTALL_MSG = """\
Claude CLI가 설치되어 있지 않습니다.

설치 방법:
  1. https://claude.ai/code 에서 Claude Code 설치
  2. 또는 터미널: npm install -g @anthropic-ai/claude-code
  3. 설치 후 'claude' 명령어로 로그인: claude login

설치 후 다시 실행해 주세요."""

CLAUDE_PROMPT = """\
당신은 한국 주식 투자 전문 재무 분석가입니다.
아래 기업 재무 데이터를 분석하고, 투자자 관점의 의견을 JSON으로만 출력하세요.
JSON 이외의 텍스트(설명, 마크다운 등)는 절대 출력하지 마세요.

{fin_table}

다음 13개 항목을 분석하여 정확히 아래 JSON 형식으로 출력하세요:
{{
  "영업이익률":     {{"value":"최신값(예:11.9%)", "text":"투자관점 의견 25자이내", "level":"good|note|warn|danger"}},
  "순이익률":       {{"value":"...", "text":"...", "level":"..."}},
  "매출총이익률":   {{"value":"...", "text":"...", "level":"..."}},
  "ROE":           {{"value":"...", "text":"...", "level":"..."}},
  "ROA":           {{"value":"...", "text":"...", "level":"..."}},
  "부채비율":       {{"value":"...", "text":"...", "level":"..."}},
  "유동비율":       {{"value":"...", "text":"...", "level":"..."}},
  "이자보상배율":   {{"value":"...(예:5.2x)", "text":"...", "level":"..."}},
  "이익잉여금 비율":{{"value":"...", "text":"...", "level":"..."}},
  "매출 성장률":    {{"value":"▲X.X% 또는 ▼X.X%", "text":"...", "level":"..."}},
  "영업이익 성장률":{{"value":"...", "text":"...", "level":"..."}},
  "순이익 성장률":  {{"value":"...", "text":"...", "level":"..."}},
  "종합 의견":      {{"value":"A|B|C|D", "text":"종합 등급 한 줄 평가", "level":"..."}}
}}

level 기준 — good:양호/강점, note:보통/중립, warn:주의/약점, danger:위험/심각"""


def check_claude_cli() -> bool:
    """Return True if 'claude' is on PATH."""
    return shutil.which("claude") is not None


def _build_fin_table(corp_name: str, yearly: dict) -> str:
    """Format financial data as a readable table for the Claude prompt."""
    avail = [yr for yr in YEARS if yearly.get(yr)]
    if not avail:
        return f"회사명: {corp_name}\n데이터 없음"

    latest = avail[-1]
    prev   = avail[-2] if len(avail) >= 2 else None
    lines  = [f"회사명: {corp_name}  |  분석 기준: {latest}년 결산\n"]

    def row(label, vals, suffix="억원", fmt_fn=None):
        parts = []
        for i, (yr, v) in enumerate(zip(avail, vals)):
            if v is None:
                parts.append(f"{yr}: N/A")
                continue
            s = fmt_fn(v) if fmt_fn else f"{v/1e8:,.1f}{suffix}"
            if i > 0 and vals[i-1] is not None and vals[i-1] != 0:
                r = (v - vals[i-1]) / abs(vals[i-1]) * 100
                arrow = f"▲{r:.1f}%" if r > 0 else f"▼{abs(r):.1f}%"
                parts.append(f"{yr}: {s} ({arrow})")
            else:
                parts.append(f"{yr}: {s}")
        return f"  {label:<16} " + "  ".join(parts)

    lines.append("[ 손익계산서 ] (단위: 억원)")
    for key in INCOME_KEYS:
        if key == "EPS":
            vals = [yearly[yr].get(key) for yr in avail]
            lines.append(row(key, vals, "원",
                             lambda v: f"{int(v):,}원"))
        else:
            vals = [yearly[yr].get(key) for yr in avail]
            lines.append(row(key, vals))

    lines.append("\n[ 재무상태표 ] (단위: 억원)")
    for key in BALANCE_KEYS:
        vals = [yearly[yr].get(key) for yr in avail]
        lines.append(row(key, vals))

    lines.append("\n[ 수익성·안정성 지표 ]")
    for key in RATIO_KEYS:
        vals = [yearly[yr].get(key) for yr in avail]
        if key in XRATE_KEYS:
            fn = lambda v: f"{v:.2f}x"
        else:
            fn = lambda v: f"{v:.1f}%"
        lines.append(row(key, vals, "", fn))

    return "\n".join(lines)


def _extract_json(text: str) -> dict | None:
    """Robustly extract a JSON object from Claude's response."""
    text = text.strip()
    # 1. Direct parse
    try:
        return json.loads(text)
    except Exception:
        pass
    # 2. Markdown code block
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        try: return json.loads(m.group(1))
        except Exception: pass
    # 3. Find outermost { ... }
    start = text.find("{")
    end   = text.rfind("}")
    if start != -1 and end > start:
        try: return json.loads(text[start:end+1])
        except Exception: pass
    return None


def _normalize_opinions(raw: dict) -> dict:
    """Ensure every OPINION_KEYS label is present with correct structure."""
    result = {}
    for cat, lbl in OPINION_KEYS:
        item = raw.get(lbl, {})
        if not isinstance(item, dict):
            item = {}
        result[lbl] = {
            "category": cat,
            "value":    str(item.get("value", "N/A")),
            "text":     str(item.get("text",  "분석 데이터 없음")),
            "level":    item.get("level", "note")
                        if item.get("level") in LV_COLOR else "note",
        }
    return result


def call_claude_cli(prompt: str, timeout: int = 120) -> str | None:
    """
    Send prompt to Claude CLI via stdin and return stdout.
    Tries 'claude --print -' then 'claude -p' as fallback.
    Returns None on error.
    """
    for cmd in [["claude", "--print", "-"], ["claude", "-p"]]:
        try:
            result = subprocess.run(
                cmd,
                input=prompt,
                capture_output=True,
                text=True,
                timeout=timeout,
                encoding="utf-8",
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except subprocess.TimeoutExpired:
            return None
        except FileNotFoundError:
            return None
        except Exception:
            continue
    return None


def generate_opinions_claude(corp_name: str, yearly: dict,
                              save_json_path: str | None = None) -> dict | None:
    """
    Generate opinions for one company using Claude CLI.
    Optionally saves financial data as JSON to save_json_path before calling.
    Returns normalised opinion dict or None on failure.
    """
    if not check_claude_cli():
        return None

    # Optionally persist financial data as JSON
    if save_json_path:
        data = {
            "corp_name": corp_name,
            "generated_at": datetime.now().isoformat(),
            "yearly": {
                yr: {k: v for k, v in d.items() if k != "_fs"}
                for yr, d in yearly.items()
            },
        }
        try:
            with open(save_json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    fin_table = _build_fin_table(corp_name, yearly)
    prompt    = CLAUDE_PROMPT.format(fin_table=fin_table)
    response  = call_claude_cli(prompt)

    if not response:
        return None

    raw = _extract_json(response)
    if not raw:
        return None

    return _normalize_opinions(raw)


def opinions_from_json_file(json_path: str) -> tuple[str, dict, dict | None]:
    """
    Load a saved financial JSON and generate opinions.
    Returns (corp_name, yearly, opinions_dict).
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    corp_name = data.get("corp_name", "알 수 없는 기업")
    yearly    = data.get("yearly", {})
    opinions  = generate_opinions_claude(corp_name, yearly)
    return corp_name, yearly, opinions


# ── DART Data Layer ────────────────────────────────────────────────────────────
class DART:
    def __init__(self):
        self.corps: dict = {}
        self.ready = False

    def load(self, cb):
        try:
            r = requests.get(CORP_URL, params={"crtfc_key": API_KEY}, timeout=30)
            r.raise_for_status()
            with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                with z.open(z.namelist()[0]) as f:
                    root = ET.parse(f).getroot()
            self.corps = {}
            for item in root.findall("list"):
                nm = item.findtext("corp_name", "").strip()
                if not nm: continue
                self.corps.setdefault(nm, []).append({
                    "corp_code":  item.findtext("corp_code", "").strip(),
                    "stock_code": item.findtext("stock_code", "").strip(),
                })
            self.ready = True
            cb(True, "준비 완료")
        except Exception as e:
            cb(False, str(e))

    def search(self, q: str) -> list:
        if not self.ready: return []
        ql = q.lower()
        return [
            {"corp_name": nm, **c}
            for nm, corps in self.corps.items()
            if ql in nm.lower()
            for c in corps
        ]

    def financials(self, corp_code: str, year: str):
        for fs in ("CFS", "OFS"):
            try:
                r = requests.get(FIN_URL, params={
                    "crtfc_key": API_KEY, "corp_code": corp_code,
                    "bsns_year": year, "reprt_code": "11011", "fs_div": fs,
                }, timeout=30)
                d = r.json()
                if d.get("status") == "000" and d.get("list"):
                    return d["list"], fs
            except Exception:
                pass
        return None, None

    def extract(self, items: list) -> dict:
        """
        account_id → CODE_MAP (1순위), account_nm → NAME_MAP (2순위).
        동일 표준항목에 여러 행이 매핑되면 절대값이 큰 값 유지.
        """
        if not items: return {}
        res: dict = {}
        for item in items:
            acct_id = (item.get("account_id") or "").strip()
            acct_nm = (item.get("account_nm") or "").strip()
            canonical = CODE_MAP.get(acct_id) or NAME_MAP.get(acct_nm)
            if not canonical:
                continue
            v = _parse(item.get("thstrm_amount", ""), as_float=(canonical in EPS_KEYS))
            if v is None:
                continue
            if canonical not in res or abs(v) > abs(res[canonical]):
                res[canonical] = v
        return res

    @staticmethod
    def calc_ratios(d: dict) -> dict:
        rev   = d.get("매출")    or 0
        op    = d.get("영업이익")
        net   = d.get("당기순이익")
        asset = d.get("자산총계")  or 0
        eq    = d.get("자본총계")  or 0
        dbt   = d.get("부채총계")  or 0
        ca    = d.get("유동자산") or 0
        cl    = d.get("유동부채") or 0
        fc    = d.get("금융비용") or 0
        gp    = d.get("매출총이익") or 0
        return {
            "영업이익률":   op  / rev   * 100 if op  is not None and rev   else None,
            "순이익률":     net / rev   * 100 if net is not None and rev   else None,
            "ROE":         net / eq    * 100 if net is not None and eq    else None,
            "ROA":         net / asset * 100 if net is not None and asset else None,
            "부채비율":     dbt / eq    * 100 if eq                        else None,
            "유동비율":     ca  / cl    * 100 if cl                        else None,
            "이자보상배율": op  / fc          if op  is not None and fc    else None,
            "매출총이익률": gp  / rev   * 100 if gp  and rev               else None,
        }


# ── Company Selection Dialog ───────────────────────────────────────────────────
class SelectDialog(ctk.CTkToplevel):
    def __init__(self, parent, corps: list, co_idx: int = 0):
        super().__init__(parent)
        self.title("기업 선택")
        self.geometry("540x440")
        self.resizable(False, False)
        self.configure(fg_color=C["bg"])
        self.grab_set()
        self.result = None
        self.update_idletasks()
        px, py = parent.winfo_x(), parent.winfo_y()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        self.geometry(f"+{px+pw//2-270}+{py+ph//2-220}")

        color = CO_COLS[co_idx % len(CO_COLS)]
        bar = ctk.CTkFrame(self, fg_color=color, corner_radius=0, height=56)
        bar.pack(fill="x"); bar.pack_propagate(False)
        ctk.CTkLabel(bar, text=f"회사 {co_idx+1}  ·  분석할 기업을 선택하세요",
                     font=ctk.CTkFont(FONT, 15, "bold"),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        ctk.CTkLabel(self, text=f"총 {len(corps)}개 기업 검색됨",
                     font=ctk.CTkFont(FONT, 11), text_color=C["g1"]).pack(pady=(10,4))
        sf = ctk.CTkScrollableFrame(self, fg_color=C["bg"], corner_radius=0)
        sf.pack(fill="both", expand=True, padx=16, pady=(0,12))
        for corp in corps:
            row = ctk.CTkFrame(sf, fg_color=C["card"], corner_radius=10,
                               border_width=1, border_color=C["g3"])
            row.pack(fill="x", pady=3); row.columnconfigure(0, weight=1)
            info = ctk.CTkFrame(row, fg_color="transparent")
            info.grid(row=0, column=0, sticky="w", padx=14, pady=10)
            ctk.CTkLabel(info, text=corp["corp_name"],
                         font=ctk.CTkFont(FONT, 13, "bold"),
                         text_color=C["blk"]).pack(anchor="w")
            ctk.CTkLabel(info,
                         text=f"종목: {corp.get('stock_code') or '비상장'}  |  "
                              f"기업코드: {corp['corp_code']}",
                         font=ctk.CTkFont(FONT, 11), text_color=C["g1"]).pack(anchor="w")
            ctk.CTkButton(row, text="선택", width=64, height=32, corner_radius=8,
                          fg_color=color, hover_color=CO_DARKS[co_idx % len(CO_DARKS)],
                          font=ctk.CTkFont(FONT, 12, "bold"),
                          command=lambda c=corp: self._pick(c),
                          ).grid(row=0, column=1, padx=(8,14), pady=10)

    def _pick(self, corp):
        self.result = corp; self.destroy()


# ── Comparison Table ───────────────────────────────────────────────────────────
class ComparisonTable(ctk.CTkFrame):
    LW = 120; CW = 84; RH = 46; GH = 34; SH = 24

    def __init__(self, parent, title, keys, cos, yearly, is_ratio=False):
        super().__init__(parent, fg_color=C["card"], corner_radius=16,
                         border_width=1, border_color=C["g3"])
        active = [(i, nm) for i, nm in enumerate(cos) if nm]
        hf = ctk.CTkFrame(self, fg_color="transparent")
        hf.pack(fill="x", padx=20, pady=(14,6))
        ctk.CTkLabel(hf, text=title, font=ctk.CTkFont(FONT, 14, "bold"),
                     text_color=C["blk"], anchor="w").pack(side="left")
        ctk.CTkFrame(self, fg_color=C["g3"], height=1).pack(fill="x", padx=20, pady=(0,6))
        outer = tk.Frame(self, bg=C["card"])
        outer.pack(fill="both", expand=True, padx=12, pady=(0,12))
        n_cols    = self.LW + len(active) * 3 * self.CW
        n_rows_px = self.GH + self.SH + len(keys) * self.RH + 4
        canvas = tk.Canvas(outer, bg=C["card"], highlightthickness=0,
                           height=min(n_rows_px, 560))
        hbar   = tk.Scrollbar(outer, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=hbar.set)
        inner = tk.Frame(canvas, bg=C["card"])
        wid   = canvas.create_window(0, 0, anchor="nw", window=inner)
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: (
            canvas.itemconfig(wid, height=max(e.height, n_rows_px)),
            canvas.itemconfig(wid, width=max(e.width, n_cols)),
        ))
        if n_cols > 900: hbar.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)
        self._paint(inner, active, keys, yearly, is_ratio)

    def _paint(self, parent, active, keys, yearly, is_ratio):
        def col(ci, yi): return 1 + ci * 3 + yi
        def cell(r, c, text, bg, fg, fsz=10, bold=False,
                 anchor="center", w=None, h=None):
            fw, fh = w or self.CW, h or self.RH
            f = tk.Frame(parent, bg=bg, width=fw, height=fh)
            f.grid(row=r, column=c, sticky="nsew"); f.grid_propagate(False)
            tk.Label(f, text=text, bg=bg, fg=fg, anchor=anchor,
                     font=(FONT, fsz, "bold" if bold else "normal"),
                     justify="center", wraplength=fw-6,
                     ).place(relx=.5, rely=.5, anchor="center")

        cell(0, 0, "", C["blk"], "white", w=self.LW, h=self.GH)
        for li, (co_i, name) in enumerate(active):
            bg = CO_COLS[co_i]
            for yi in range(3):
                f = tk.Frame(parent, bg=bg, width=self.CW, height=self.GH)
                f.grid(row=0, column=col(li, yi), sticky="nsew"); f.grid_propagate(False)
                if yi == 1:
                    tk.Label(f, text=name, bg=bg, fg="white",
                             font=(FONT, 11, "bold")).place(relx=.5, rely=.5, anchor="center")

        cell(1, 0, "항목", "#2D3540", C["g2"], 10, anchor="w", w=self.LW, h=self.SH)
        for li, (co_i, _) in enumerate(active):
            for yi, yr in enumerate(YEARS):
                cell(1, col(li, yi), yr, "#2D3540", C["g2"], 9, w=self.CW, h=self.SH)

        for ri, key in enumerate(keys):
            bg = C["g4"] if ri % 2 == 0 else C["card"]
            r  = ri + 2
            cell(r, 0, key, bg, C["g1"], 11, anchor="w", w=self.LW, h=self.RH)
            for li, (co_i, _) in enumerate(active):
                yr_vals = [
                    yearly[co_i].get(yr, {}).get(key)
                    if yearly[co_i].get(yr) else None for yr in YEARS
                ]
                for yi, v in enumerate(yr_vals):
                    c  = col(li, yi)
                    vs = fmt_val(v, key, is_ratio)
                    fg = C["blk"] if v is not None else C["g2"]
                    f  = tk.Frame(parent, bg=bg, width=self.CW, height=self.RH)
                    f.grid(row=r, column=c, sticky="nsew"); f.grid_propagate(False)
                    tk.Label(f, text=vs, bg=bg, fg=fg,
                             font=(FONT, 11, "bold")).place(x=self.CW//2, y=12, anchor="center")
                    if yi > 0:
                        txt, col_fg = yoy(v, yr_vals[yi-1])
                        if txt:
                            tk.Label(f, text=txt, bg=bg, fg=col_fg or C["g1"],
                                     font=(FONT, 9)).place(x=self.CW//2, y=30, anchor="center")


# ── Opinion Panel ──────────────────────────────────────────────────────────────
class OpinionPanel(ctk.CTkFrame):
    """3-company side-by-side Claude AI opinion card."""

    def __init__(self, parent, cos: list, opinions: list,
                 claude_available: bool = True):
        super().__init__(parent, fg_color=C["card"], corner_radius=16,
                         border_width=1, border_color=C["g3"])
        active = [(i, nm) for i, nm in enumerate(cos) if nm]

        # Header
        hf = ctk.CTkFrame(self, fg_color="transparent")
        hf.pack(fill="x", padx=20, pady=(14,6))
        ctk.CTkLabel(hf, text="재무 분석 의견",
                     font=ctk.CTkFont(FONT, 14, "bold"),
                     text_color=C["blk"], anchor="w").pack(side="left")

        # Claude badge
        badge_color = C["blue"] if claude_available else C["g1"]
        badge_text  = "Claude AI" if claude_available else "Claude CLI 미설치"
        chip = ctk.CTkFrame(hf, fg_color=badge_color, corner_radius=6)
        chip.pack(side="right")
        ctk.CTkLabel(chip, text=badge_text,
                     font=ctk.CTkFont(FONT, 10, "bold"),
                     text_color="white").pack(padx=8, pady=3)

        ctk.CTkFrame(self, fg_color=C["g3"], height=1).pack(fill="x", padx=20, pady=(0,8))

        # If Claude not available
        if not claude_available:
            info = ctk.CTkFrame(self, fg_color=LV_BG["warn"], corner_radius=12)
            info.pack(fill="x", padx=16, pady=(0,16))
            ctk.CTkLabel(info, text=CLAUDE_INSTALL_MSG,
                         font=ctk.CTkFont(FONT, 11),
                         text_color=LV_COLOR["warn"],
                         justify="left", anchor="w").pack(padx=20, pady=16, anchor="w")
            return

        # Column headers
        ch = ctk.CTkFrame(self, fg_color="transparent")
        ch.pack(fill="x", padx=16, pady=(0,6))
        ctk.CTkLabel(ch, text="", width=186).pack(side="left")
        for ci, nm in active:
            chip2 = ctk.CTkFrame(ch, fg_color=CO_COLS[ci], corner_radius=8)
            chip2.pack(side="left", padx=4, expand=True, fill="x")
            ctk.CTkLabel(chip2, text=f" {nm[:16]} ",
                         font=ctk.CTkFont(FONT, 11, "bold"),
                         text_color="white").pack(padx=6, pady=5)

        # Opinion rows
        for row_i, (cat, lbl) in enumerate(OPINION_KEYS):
            bg = C["g4"] if row_i % 2 == 0 else C["card"]
            rf = ctk.CTkFrame(self, fg_color=bg, corner_radius=8)
            rf.pack(fill="x", padx=12, pady=1)

            label_f = ctk.CTkFrame(rf, fg_color="transparent", width=186)
            label_f.pack(side="left", padx=(8,4), pady=6); label_f.pack_propagate(False)
            cat_chip = ctk.CTkFrame(label_f, fg_color=C["g3"], corner_radius=4)
            cat_chip.pack(anchor="w")
            ctk.CTkLabel(cat_chip, text=cat, font=ctk.CTkFont(FONT, 9),
                         text_color=C["g1"]).pack(padx=6, pady=1)
            ctk.CTkLabel(label_f, text=lbl,
                         font=ctk.CTkFont(FONT, 12, "bold"),
                         text_color=C["blk"], anchor="w").pack(anchor="w", pady=(2,0))

            for ci, nm in active:
                op   = opinions[ci] or {}
                item = op.get(lbl, {"value":"…", "text":"생성 중 또는 실패", "level":"note"})
                lv   = item.get("level", "note")
                fg   = LV_COLOR.get(lv, C["g1"])
                bg2  = LV_BG.get(lv, C["g4"])
                icon = LV_ICON.get(lv, "·")

                cell_f = ctk.CTkFrame(rf, fg_color=bg2, corner_radius=8)
                cell_f.pack(side="left", padx=4, pady=4, expand=True, fill="both")

                top = ctk.CTkFrame(cell_f, fg_color="transparent")
                top.pack(fill="x", padx=8, pady=(4,0))
                ctk.CTkLabel(top, text=f"{icon}  {item.get('value','N/A')}",
                             font=ctk.CTkFont(FONT, 12, "bold"),
                             text_color=fg).pack(side="left")

                ctk.CTkLabel(cell_f, text=item.get("text",""),
                             font=ctk.CTkFont(FONT, 10), text_color=C["g1"],
                             anchor="w", wraplength=260, justify="left",
                             ).pack(fill="x", padx=8, pady=(2,6))

        ctk.CTkFrame(self, fg_color="transparent", height=8).pack()


# ── Search Row ─────────────────────────────────────────────────────────────────
class SearchRow(ctk.CTkFrame):
    def __init__(self, parent, idx: int, on_search):
        super().__init__(parent, fg_color="transparent")
        self.idx = idx; self.on_search = on_search; self.selected = None
        self._build()

    def _build(self):
        color = CO_COLS[self.idx]
        badge = ctk.CTkFrame(self, width=28, height=28, corner_radius=14, fg_color=color)
        badge.pack(side="left", padx=(0,8)); badge.pack_propagate(False)
        ctk.CTkLabel(badge, text=str(self.idx+1),
                     font=ctk.CTkFont(FONT, 12, "bold"),
                     text_color="white").place(relx=.5, rely=.5, anchor="center")
        self.entry = ctk.CTkEntry(
            self, placeholder_text=f"회사 {self.idx+1} 기업명",
            width=260, height=40, corner_radius=10,
            border_color=C["g3"], fg_color=C["g4"],
            text_color=C["blk"], font=ctk.CTkFont(FONT, 12))
        self.entry.pack(side="left", padx=(0,8))
        self.entry.bind("<Return>", lambda _: self._do_search())
        ctk.CTkButton(self, text="검색", width=64, height=40, corner_radius=10,
                      fg_color=color, hover_color=CO_DARKS[self.idx],
                      font=ctk.CTkFont(FONT, 12, "bold"),
                      command=self._do_search).pack(side="left", padx=(0,10))
        self.lbl = ctk.CTkLabel(self, text="미선택",
                                font=ctk.CTkFont(FONT, 12),
                                text_color=C["g2"], width=160, anchor="w")
        self.lbl.pack(side="left")
        self.clear_btn = ctk.CTkButton(
            self, text="✕", width=28, height=28, corner_radius=8,
            fg_color=C["g3"], hover_color=C["g2"], text_color=C["g1"],
            font=ctk.CTkFont(FONT, 11), command=self.clear)

    def _do_search(self):
        self.on_search(self.idx, self.entry.get().strip())

    def set_company(self, corp: dict):
        self.selected = corp
        nm = corp["corp_name"]
        self.lbl.configure(
            text=f"✓  {nm[:14]}{'…' if len(nm)>14 else ''}",
            text_color=CO_COLS[self.idx])
        self.clear_btn.pack(side="left", padx=(4,0))

    def clear(self):
        self.selected = None
        self.lbl.configure(text="미선택", text_color=C["g2"])
        self.entry.delete(0, "end")
        self.clear_btn.pack_forget()


# ── Main Application ───────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("기업 재무분석기 v2  ·  3사 비교 + Claude AI")
        self.geometry("1280x920")
        self.minsize(1000, 720)
        self.configure(fg_color=C["bg"])
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.dart             = DART()
        self.yearly           = [{}, {}, {}]
        self.cos              = [None, None, None]
        self.opinions         = [None, None, None]
        self._widgets: list   = []
        self._loading_lbl     = None
        self._claude_ok       = check_claude_cli()

        self._build_ui()
        threading.Thread(target=self.dart.load, args=(self._on_load,),
                         daemon=True).start()

    # ── UI ─────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        nav = ctk.CTkFrame(self, fg_color=C["card"], corner_radius=0, height=60)
        nav.pack(fill="x"); nav.pack_propagate(False)
        logo = ctk.CTkFrame(nav, fg_color="transparent")
        logo.place(relx=.5, rely=.5, anchor="center")
        ctk.CTkFrame(logo, width=8, height=8, corner_radius=4,
                     fg_color=C["blue"]).pack(side="left", padx=(0,8))
        ctk.CTkLabel(logo, text="기업 재무분석기",
                     font=ctk.CTkFont(FONT, 18, "bold"),
                     text_color=C["blk"]).pack(side="left")

        # Claude CLI status badge
        cl_color = C["blue"] if self._claude_ok else C["g2"]
        cl_text  = "Claude CLI ✓" if self._claude_ok else "Claude CLI 미설치"
        cl_chip  = ctk.CTkFrame(logo, fg_color=cl_color, corner_radius=6)
        cl_chip.pack(side="left", padx=(10,0))
        ctk.CTkLabel(cl_chip, text=cl_text,
                     font=ctk.CTkFont(FONT, 10, "bold"),
                     text_color="white").pack(padx=8, pady=3)

        self.status_lbl = ctk.CTkLabel(nav, text="기업 코드 로딩 중…",
                                        font=ctk.CTkFont(FONT, 11), text_color=C["g1"])
        self.status_lbl.place(relx=.98, rely=.5, anchor="e")

        sp = ctk.CTkFrame(self, fg_color=C["card"], corner_radius=0)
        sp.pack(fill="x")
        rows_frame = ctk.CTkFrame(sp, fg_color="transparent")
        rows_frame.pack(anchor="center", padx=40, pady=(14,8))
        self.search_rows: list[SearchRow] = []
        for i in range(3):
            sr = SearchRow(rows_frame, i, self._on_search_btn)
            sr.pack(fill="x", pady=4)
            self.search_rows.append(sr)

        btn_row = ctk.CTkFrame(sp, fg_color="transparent")
        btn_row.pack(pady=(4,14))
        self.btn_run = ctk.CTkButton(
            btn_row, text="비교 분석 시작", width=160, height=44,
            corner_radius=12, fg_color=C["blue"], hover_color=C["bdark"],
            font=ctk.CTkFont(FONT, 13, "bold"), command=self._run_analysis)
        self.btn_run.pack(side="left", padx=(0,10))
        self.btn_excel = ctk.CTkButton(
            btn_row, text="⬇  Excel 저장", width=140, height=44,
            corner_radius=12, fg_color=C["grn"], hover_color="#04A862",
            font=ctk.CTkFont(FONT, 13, "bold"),
            command=self._save_excel, state="disabled")
        self.btn_excel.pack(side="left")

        ctk.CTkFrame(self, fg_color=C["g3"], height=1).pack(fill="x")
        self.scroll = ctk.CTkScrollableFrame(
            self, fg_color=C["bg"], corner_radius=0,
            scrollbar_button_color=C["g3"],
            scrollbar_button_hover_color=C["g2"])
        self.scroll.pack(fill="both", expand=True)
        self._show_placeholder()

    def _show_placeholder(self):
        ph = ctk.CTkFrame(self.scroll, fg_color="transparent")
        ph.pack(pady=80); self._widgets.append(ph)
        ctk.CTkLabel(ph, text="🏢", font=ctk.CTkFont(size=52)).pack()
        ctk.CTkLabel(ph, text="최대 3개 기업을 입력하고 비교 분석을 시작하세요",
                     font=ctk.CTkFont(FONT, 15), text_color=C["g2"]).pack(pady=(14,0))
        hint = "26개 계정  ·  3개년  ·  YoY  ·  Claude AI 의견  ·  Excel"
        if not self._claude_ok:
            hint += "\n⚠ Claude CLI 미설치 — 의견 기능 비활성화"
        ctk.CTkLabel(ph, text=hint, font=ctk.CTkFont(FONT, 12),
                     text_color=C["g2"] if self._claude_ok else LV_COLOR["warn"],
                     justify="center").pack(pady=(6,0))

    # ── Callbacks ──────────────────────────────────────────────────────────────
    def _on_load(self, ok, msg=""):
        def _u():
            if ok:  self.status_lbl.configure(text="✓ 준비 완료", text_color=C["grn"])
            else:   self.status_lbl.configure(text="✗ 로드 실패", text_color=C["red"])
        self.after(0, _u)

    def _on_search_btn(self, idx, q):
        if not q: return
        if not self.dart.ready:
            self._toast("기업 코드 로딩 중입니다."); return
        corps = self.dart.search(q)
        if not corps:
            self._toast(f"'{q}' 기업을 찾을 수 없습니다."); return
        exact = [c for c in corps if c["corp_name"] == q]
        pool  = exact if exact else corps
        if len(pool) == 1:
            self.search_rows[idx].set_company(pool[0])
        else:
            dlg = SelectDialog(self, pool[:30], idx)
            self.wait_window(dlg)
            if dlg.result: self.search_rows[idx].set_company(dlg.result)

    def _run_analysis(self):
        if not any(sr.selected for sr in self.search_rows):
            self._toast("최소 1개 기업을 선택해 주세요."); return
        self._clear()
        self._show_loading()
        self.cos = [sr.selected for sr in self.search_rows]
        threading.Thread(target=self._fetch_all, daemon=True).start()

    def _update_loading(self, text: str):
        """Thread-safe loading label update."""
        def _u():
            if self._loading_lbl:
                try: self._loading_lbl.configure(text=text)
                except Exception: pass
        self.after(0, _u)

    def _fetch_all(self):
        # ── Phase 1: DART — account_id 기반 재무 추출 ─────────────────────────
        yearly = [{}, {}, {}]
        for ci, corp in enumerate(self.cos):
            if not corp: continue
            nm = corp["corp_name"]
            self._update_loading(f"DART 조회 중: {nm}…")
            for yr in YEARS:
                items, fs = self.dart.financials(corp["corp_code"], yr)
                if items:
                    fin = self.dart.extract(items)
                    rat = DART.calc_ratios(fin)
                    yearly[ci][yr] = {**fin, **rat, "_fs": fs}
        self.yearly = yearly

        # ── Phase 2: Claude CLI — 재무 의견 생성 (Sheet5) ─────────────────────
        opinions = [None, None, None]
        if not self._claude_ok:
            self._update_loading("Claude CLI 미설치 — 의견 생성 건너뜀")
        else:
            tmpdir = tempfile.gettempdir()
            for ci, corp in enumerate(self.cos):
                if not corp: continue
                nm = corp["corp_name"]
                self._update_loading(f"Claude AI 분석 중: {nm}…  (최대 2분 소요)")
                json_path = os.path.join(tmpdir, f"dart_fin_{ci}_{nm[:8]}.json")
                op = generate_opinions_claude(nm, yearly[ci], json_path)
                opinions[ci] = op
                if op is None:
                    self._update_loading(f"⚠ {nm} Claude 응답 없음 — 재시도 없이 건너뜀")

        self.opinions = opinions
        self.after(0, lambda: self._display(yearly))

    # ── Display ────────────────────────────────────────────────────────────────
    def _clear(self):
        for w in self._widgets:
            try: w.destroy()
            except Exception: pass
        self._widgets.clear()
        self._loading_lbl = None

    def _show_loading(self):
        f = ctk.CTkFrame(self.scroll, fg_color=C["card"], corner_radius=16)
        f.pack(fill="x", padx=40, pady=40); self._widgets.append(f)
        self._loading_lbl = ctk.CTkLabel(
            f, text="DART 재무 데이터 조회 중…",
            font=ctk.CTkFont(FONT, 14), text_color=C["g1"])
        self._loading_lbl.pack(pady=(32,12))
        pb = ctk.CTkProgressBar(f, mode="indeterminate", width=320,
                                fg_color=C["g3"], progress_color=C["blue"])
        pb.pack(pady=(0,8)); pb.start()
        ctk.CTkLabel(
            f, text="① DART 재무 조회 (account_id 기반)  →  ② Claude AI 의견 생성 (Sheet5)",
            font=ctk.CTkFont(FONT, 11), text_color=C["g2"],
        ).pack(pady=(0,32))

    def _display(self, yearly):
        self._clear()
        co_names = [sr.selected["corp_name"] if sr.selected else None
                    for sr in self.search_rows]
        active   = [(i, nm) for i, nm in enumerate(co_names) if nm]

        banner = ctk.CTkFrame(self.scroll, fg_color=C["blk"], corner_radius=16)
        banner.pack(fill="x", padx=40, pady=(24,10)); self._widgets.append(banner)
        bi = ctk.CTkFrame(banner, fg_color="transparent")
        bi.pack(fill="x", padx=24, pady=16)
        for i, (co_i, name) in enumerate(active):
            chip = ctk.CTkFrame(bi, fg_color=CO_COLS[co_i], corner_radius=8)
            chip.pack(side="left", padx=(0 if i == 0 else 8, 0))
            ctk.CTkLabel(chip, text=f" {i+1}  {name} ",
                         font=ctk.CTkFont(FONT, 13, "bold"),
                         text_color="white").pack(padx=8, pady=6)
        ctk.CTkLabel(bi, text="26개 계정  ·  3개년  ·  Claude AI 의견",
                     font=ctk.CTkFont(FONT, 11),
                     text_color="#FFFFFF66").pack(side="right")

        for title, keys, is_ratio in SECTIONS:
            tbl = ComparisonTable(self.scroll, title, keys, co_names, yearly, is_ratio)
            tbl.pack(fill="x", padx=40, pady=8); self._widgets.append(tbl)

        op_panel = OpinionPanel(
            self.scroll, co_names, self.opinions, self._claude_ok)
        op_panel.pack(fill="x", padx=40, pady=(8,30))
        self._widgets.append(op_panel)

        self.btn_excel.configure(state="normal")

    # ── Excel ──────────────────────────────────────────────────────────────────
    def _save_excel(self):
        if not any(c for c in self.cos): return
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        names = [c["corp_name"] for c in self.cos if c]
        stem  = "_".join(names[:2]) + ("_외" if len(names) > 2 else "")
        path  = os.path.join(os.path.expanduser("~/Desktop"),
                             f"{stem}_재무비교_{ts}.xlsx")
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        for ci, corp in enumerate(self.cos):
            if not corp: continue
            ws = wb.create_sheet(title=corp["corp_name"][:28])
            self._write_company_sheet(ws, ci, corp["corp_name"])
        ws_cmp = wb.create_sheet(title="전체 비교")
        self._write_comparison_sheet(ws_cmp)
        ws_op = wb.create_sheet(title="재무분석 의견")
        self._write_opinion_sheet(ws_op)
        # Reorder sheets
        for target in reversed(["재무분석 의견", "전체 비교"]):
            if target in wb.sheetnames:
                wb.move_sheet(target, offset=-len(wb.sheetnames))
        wb.save(path)
        self._toast(f"저장 완료 → {os.path.basename(path)}")

    def _xl_styles(self):
        def mf(bold=False, color="191F28", size=11, italic=False):
            return Font(name="맑은 고딕", bold=bold, color=color, size=size, italic=italic)
        def fl(hex_color):
            return PatternFill("solid", fgColor=hex_color.lstrip("#"))
        BD = Border(
            left=Side(style="thin", color="E5E8EB"),
            right=Side(style="thin", color="E5E8EB"),
            top=Side(style="thin", color="E5E8EB"),
            bottom=Side(style="thin", color="E5E8EB"),
        )
        CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        return mf, fl, BD, CTR, LEFT

    def _write_company_sheet(self, ws, ci, corp_name):
        mf, fl, BD, CTR, LEFT = self._xl_styles()
        co_hex = CO_COLS[ci].lstrip("#")
        def cell(r, c, v="", font=None, bg=None, align=CTR):
            cc = ws.cell(r, c, v)
            if font: cc.font = font
            if bg:   cc.fill = fl(bg)
            cc.alignment = align; cc.border = BD
        ws.merge_cells("A1:G2")
        cell(1,1,f"{corp_name}  재무분석",mf(True,"FFFFFF",14),co_hex,CTR)
        ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
        ws.merge_cells("A3:G3")
        cell(3,1,f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
             mf(color="6B7684",size=10,italic=True),"FFFFFF",LEFT)
        ws.row_dimensions[3].height = 16
        r = 5
        for sec_title, keys, is_ratio in SECTIONS:
            ws.merge_cells(f"A{r}:G{r}")
            cell(r,1,sec_title,mf(True,"FFFFFF",12),co_hex,CTR)
            ws.row_dimensions[r].height = 26; r += 1
            for ci2,h in enumerate(["항목","2023","2024","YoY","2025","YoY",""],1):
                cell(r,ci2,h,mf(True,"FFFFFF",10),"6B7684",CTR)
            ws.row_dimensions[r].height = 20; r += 1
            for ki, key in enumerate(keys):
                vals = [self.yearly[ci].get(yr,{}).get(key)
                        if self.yearly[ci].get(yr) else None for yr in YEARS]
                bg = "F9FAFB" if ki%2==0 else "FFFFFF"
                ws.row_dimensions[r].height = 22
                cell(r,1,key,mf(color="6B7684"),bg,LEFT)
                cell(r,2,fmt_val(vals[0],key,is_ratio),mf(True),bg,CTR)
                cell(r,7,"",mf(),bg,CTR)
                for col_off, idx in [(3,1),(5,2)]:
                    v = vals[idx]
                    cell(r,col_off,fmt_val(v,key,is_ratio),mf(True),bg,CTR)
                    txt,col_c = yoy(v,vals[idx-1])
                    if col_c==C["grn"]:   yf=mf(color="05C072",size=10)
                    elif col_c==C["red"]: yf=mf(color="F04452",size=10)
                    else:                 yf=mf(color="B0B8C1",size=10)
                    cell(r,col_off+1,txt or "-",yf,bg,CTR)
                r += 1
            r += 1
        for col_l,w in [("A",14),("B",14),("C",14),("D",11),
                        ("E",14),("F",11),("G",6)]:
            ws.column_dimensions[col_l].width = w

    def _write_comparison_sheet(self, ws):
        mf, fl, BD, CTR, LEFT = self._xl_styles()
        active   = [(ci, c["corp_name"]) for ci, c in enumerate(self.cos) if c]
        n        = len(active)
        last_col = 1 + n * 3 + 1
        def cell(r, c, v="", font=None, bg=None, align=CTR):
            cc = ws.cell(r, c, v)
            if font: cc.font = font
            if bg:   cc.fill = fl(bg)
            cc.alignment = align; cc.border = BD
        ws.merge_cells(f"A1:{chr(64+last_col)}2")
        cell(1,1,"3사 재무 비교 분석",mf(True,"FFFFFF",14),"191F28",CTR)
        ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
        ws.merge_cells(f"A3:{chr(64+last_col)}3")
        cell(3,1,f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  DART",
             mf(color="6B7684",size=10,italic=True),"FFFFFF",LEFT)
        ws.row_dimensions[3].height = 16
        r = 5
        for sec_title, keys, is_ratio in SECTIONS:
            ws.merge_cells(f"A{r}:{chr(64+last_col)}{r}")
            cell(r,1,sec_title,mf(True,"FFFFFF",12),"4B9FFF",CTR)
            ws.row_dimensions[r].height = 26; r += 1
            cell(r,1,"항목",mf(True,"FFFFFF",10),"6B7684",LEFT)
            for li,(ci,nm) in enumerate(active):
                co_hex = CO_COLS[ci].lstrip("#")
                base_c = 2 + li*3
                ws.merge_cells(f"{chr(64+base_c)}{r}:{chr(64+base_c+2)}{r}")
                cell(r,base_c,nm,mf(True,"FFFFFF",11),co_hex,CTR)
            ws.row_dimensions[r].height = 26; r += 1
            cell(r,1,"",mf(True,"FFFFFF",9),"6B7684",CTR)
            for li,(ci,_) in enumerate(active):
                for yi,yr in enumerate(YEARS):
                    cell(r,2+li*3+yi,yr,mf(True,"FFFFFF",9),"6B7684",CTR)
            ws.row_dimensions[r].height = 18; r += 1
            for ki,key in enumerate(keys):
                bg = "F9FAFB" if ki%2==0 else "FFFFFF"
                ws.row_dimensions[r].height = 22
                cell(r,1,key,mf(color="6B7684"),bg,LEFT)
                for li,(ci,_) in enumerate(active):
                    for yi,yr in enumerate(YEARS):
                        v = (self.yearly[ci].get(yr,{}).get(key)
                             if self.yearly[ci].get(yr) else None)
                        cell(r,2+li*3+yi,fmt_val(v,key,is_ratio),mf(True),bg,CTR)
                r += 1
            r += 1
        ws.column_dimensions["A"].width = 14
        for li in range(n):
            for yi in range(3):
                ws.column_dimensions[chr(66+li*3+yi)].width = 13

    def _write_opinion_sheet(self, ws):
        mf, fl, BD, CTR, LEFT = self._xl_styles()
        active = [(ci, c["corp_name"]) for ci, c in enumerate(self.cos) if c]
        n = len(active)
        def cell(r, c, v="", font=None, bg=None, align=CTR):
            cc = ws.cell(r, c, v)
            if font: cc.font = font
            if bg:   cc.fill = fl(bg)
            cc.alignment = align; cc.border = BD
        last_c = 3 + n
        ws.merge_cells(f"A1:{chr(64+last_c)}2")
        cell(1,1,"재무 분석 의견 — Claude AI 자동 평가",
             mf(True,"FFFFFF",14),"191F28",CTR)
        ws.row_dimensions[1].height = ws.row_dimensions[2].height = 18
        ws.merge_cells(f"A3:{chr(64+last_c)}3")
        cell(3,1,
             f"생성일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  Claude CLI  |  최신 결산 기준",
             mf(color="6B7684",size=10,italic=True),"FFFFFF",LEFT)
        ws.row_dimensions[3].height = 16
        r = 5
        cell(r,1,"분류",mf(True,"FFFFFF",10),"6B7684",CTR)
        cell(r,2,"평가 항목",mf(True,"FFFFFF",10),"6B7684",CTR)
        for li,(ci,nm) in enumerate(active):
            co_hex = CO_COLS[ci].lstrip("#")
            cell(r,3+li,nm,mf(True,"FFFFFF",11),co_hex,CTR)
        ws.row_dimensions[r].height = 22; r += 1
        prev_cat = None
        for row_i,(cat,lbl) in enumerate(OPINION_KEYS):
            bg_lbl = "EFF6FF" if cat != prev_cat else "FFFFFF"
            prev_cat = cat
            ws.row_dimensions[r].height = 54
            cell(r,1,cat,mf(True,"6B7684",10),bg_lbl,CTR)
            cell(r,2,lbl,mf(bold=True,size=11),"FFFFFF",LEFT)
            for li,(ci,nm) in enumerate(active):
                op   = self.opinions[ci] or {}
                item = op.get(lbl, {"value":"N/A","text":"데이터 없음","level":"note"})
                lv   = item.get("level","note")
                icon = LV_ICON.get(lv,"·")
                fg   = LV_XL.get(lv,"B0B8C1")
                bg2  = LV_XL_BG.get(lv,"FFFFFF")
                combined = f"{icon} {item.get('value','N/A')}\n{item.get('text','')}"
                cc   = ws.cell(r, 3+li, combined)
                cc.font      = Font(name="맑은 고딕", size=10, color=fg,
                                    bold=(lv=="good"))
                cc.fill      = PatternFill("solid", fgColor=bg2)
                cc.alignment = Alignment(horizontal="left", vertical="center",
                                         wrap_text=True)
                cc.border    = BD
            r += 1
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        for li in range(n):
            ws.column_dimensions[chr(67+li)].width = 44

    # ── Toast ───────────────────────────────────────────────────────────────────
    def _toast(self, msg):
        t = ctk.CTkToplevel(self)
        t.overrideredirect(True); t.attributes("-topmost", True)
        f = ctk.CTkFrame(t, fg_color=C["blk"], corner_radius=12); f.pack()
        ctk.CTkLabel(f, text=msg, font=ctk.CTkFont(FONT, 12),
                     text_color="white").pack(padx=22, pady=13)
        self.update_idletasks()
        x = self.winfo_x() + self.winfo_width()  // 2 - 180
        y = self.winfo_y() + self.winfo_height() -  80
        t.geometry(f"+{x}+{y}")
        self.after(2800, t.destroy)


# ── CLI Mode ───────────────────────────────────────────────────────────────────
def _cli_main(json_path: str):
    """Non-GUI CLI mode: python script.py data.json"""

    if not os.path.exists(json_path):
        print(f"[오류] 파일을 찾을 수 없습니다: {json_path}")
        sys.exit(1)

    if not check_claude_cli():
        print("=" * 60)
        print(CLAUDE_INSTALL_MSG)
        print("=" * 60)
        sys.exit(1)

    print(f"\n파일 로드: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    corp_name = data.get("corp_name", "알 수 없는 기업")
    yearly    = data.get("yearly", {})
    avail     = [yr for yr in YEARS if yearly.get(yr)]

    if not avail:
        print(f"[오류] {corp_name}: 조회 가능한 연도 데이터가 없습니다.")
        sys.exit(1)

    print(f"회사명  : {corp_name}")
    print(f"데이터  : {', '.join(avail)}년")
    print(f"\nClaude AI 분석 시작…  (최대 2분 소요)\n")

    opinions = generate_opinions_claude(corp_name, yearly)

    if opinions is None:
        print("[오류] Claude CLI 응답을 받지 못했습니다.")
        print("  · claude login 완료 여부 확인")
        print("  · 네트워크 연결 확인")
        sys.exit(1)

    # Pretty print
    SEP = "─" * 62
    print(SEP)
    print(f"  {corp_name}  재무 분석 의견  ({avail[-1]}년 기준)")
    print(SEP)
    prev_cat = None
    for cat, lbl in OPINION_KEYS:
        if cat != prev_cat:
            print(f"\n  ▶ {cat}")
            prev_cat = cat
        item = opinions.get(lbl, {})
        lv   = item.get("level", "note")
        icon = LV_ICON.get(lv, "·")
        val  = item.get("value", "N/A")
        txt  = item.get("text", "")
        print(f"    {icon}  {lbl:<16}  {val:<10}  {txt}")
    print(f"\n{SEP}\n")

    # Also save result JSON alongside the input file
    out_path = json_path.replace(".json", "_opinion.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"corp_name": corp_name,
                   "generated_at": datetime.now().isoformat(),
                   "opinions": opinions}, f, ensure_ascii=False, indent=2)
    print(f"결과 저장: {out_path}\n")


# ── Entry ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) > 1:
        _cli_main(sys.argv[1])
    else:
        app = App()
        app.mainloop()
