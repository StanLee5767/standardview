#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import customtkinter as ctk
from dotenv import load_dotenv
import requests
import zipfile
import xml.etree.ElementTree as ET
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading
from datetime import datetime

# ── Config ───────────────────────────────────────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
API_KEY  = os.getenv("DART_API_KEY", "")
CORP_URL = "https://opendart.fss.or.kr/api/corpCode.xml"
FIN_URL  = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
YEARS    = ["2023", "2024", "2025"]

# Toss-style palette
C = {
    "bg":    "#F2F4F6",
    "card":  "#FFFFFF",
    "blue":  "#3182F6",
    "bdark": "#1B64DA",
    "green": "#05C072",
    "red":   "#F04452",
    "black": "#191F28",
    "g1":    "#6B7684",
    "g2":    "#B0B8C1",
    "g3":    "#E5E8EB",
    "g4":    "#F9FAFB",
}
FONT = "Apple SD Gothic Neo"

# DART XBRL account mappings
ACCTS = {
    "매출액": (
        ["ifrs-full_Revenue", "dart_Revenue",
         "ifrs-full_RevenueFromContractsWithCustomers",
         "ifrs_Revenue"],
        ["매출액", "수익(매출액)", "영업수익", "매출"],
    ),
    "영업이익": (
        ["dart_OperatingIncomeLoss",
         "ifrs-full_ProfitLossFromOperatingActivities"],
        ["영업이익", "영업이익(손실)"],
    ),
    "당기순이익": (
        ["ifrs-full_ProfitLoss", "us-gaap_NetIncomeLoss",
         "dart_ProfitLoss"],
        ["당기순이익", "당기순이익(손실)", "분기순이익"],
    ),
    "자산총계": (
        ["ifrs-full_Assets", "us-gaap_Assets"],
        ["자산총계", "자산 합계"],
    ),
    "부채총계": (
        ["ifrs-full_Liabilities", "us-gaap_Liabilities"],
        ["부채총계", "부채 합계"],
    ),
    "자본총계": (
        ["ifrs-full_Equity", "us-gaap_StockholdersEquity",
         "ifrs-full_EquityAttributableToOwnersOfParent"],
        ["자본총계", "자본 합계"],
    ),
}

# ── Utilities ─────────────────────────────────────────────────────────────────
def _parse(s):
    try:
        s = str(s).replace(",", "").strip()
        if s and s not in ("-", ""):
            return int(s)
    except Exception:
        pass
    return None


def fmt_num(n):
    if n is None:
        return "N/A"
    v = n / 1e8
    if abs(v) >= 10000:
        return f"{v / 10000:,.1f}조"
    return f"{v:,.1f}억"


def fmt_pct(v):
    if v is None:
        return "N/A"
    return f"{v:.1f}%"


def yoy(cur, prev):
    """Returns (text, color) for YoY display."""
    if cur is None or prev is None or prev == 0:
        return "", None
    r = (cur - prev) / abs(prev) * 100
    if r > 0:
        return f"▲ {abs(r):.1f}%", C["green"]
    elif r < 0:
        return f"▼ {abs(r):.1f}%", C["red"]
    return "- 0.0%", C["g1"]


# ── DART Data Layer ───────────────────────────────────────────────────────────
class DART:
    def __init__(self):
        self.corps: dict = {}   # name -> [{corp_code, stock_code}]
        self.ready = False

    def load(self, cb):
        """Download & parse corp code XML. Calls cb(ok, msg)."""
        try:
            r = requests.get(CORP_URL, params={"crtfc_key": API_KEY}, timeout=30)
            r.raise_for_status()
            with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                with z.open(z.namelist()[0]) as f:
                    root = ET.parse(f).getroot()

            self.corps = {}
            for item in root.findall("list"):
                name = item.findtext("corp_name", "").strip()
                if not name:
                    continue
                self.corps.setdefault(name, []).append({
                    "corp_code":  item.findtext("corp_code", "").strip(),
                    "stock_code": item.findtext("stock_code", "").strip(),
                })
            self.ready = True
            cb(True, "준비 완료")
        except Exception as e:
            cb(False, str(e))

    def search(self, q: str) -> list:
        if not self.ready:
            return []
        q_lower = q.lower()
        out = []
        for name, corps in self.corps.items():
            if q_lower in name.lower():
                for c in corps:
                    out.append({"corp_name": name, **c})
        return out

    def financials(self, corp_code: str, year: str):
        """Fetch annual financial statements. Returns (items, fs_div) or (None, None)."""
        for fs in ("CFS", "OFS"):
            try:
                r = requests.get(FIN_URL, params={
                    "crtfc_key":  API_KEY,
                    "corp_code":  corp_code,
                    "bsns_year":  year,
                    "reprt_code": "11011",  # 사업보고서
                    "fs_div":     fs,
                }, timeout=30)
                d = r.json()
                if d.get("status") == "000" and d.get("list"):
                    return d["list"], fs
            except Exception:
                pass
        return None, None

    def extract(self, items: list) -> dict:
        """Extract target accounts from raw API list."""
        if not items:
            return {}
        res = {}
        for key, (ids, names) in ACCTS.items():
            # 1st pass: by account_id
            for item in items:
                if item.get("account_id", "") in ids:
                    v = _parse(item.get("thstrm_amount", ""))
                    if v is not None:
                        res[key] = v
                        break
            if key in res:
                continue
            # 2nd pass: by account_nm
            for item in items:
                if item.get("account_nm", "").strip() in names:
                    v = _parse(item.get("thstrm_amount", ""))
                    if v is not None:
                        res[key] = v
                        break
        return res

    @staticmethod
    def ratios(d: dict) -> dict:
        rev = d.get("매출액") or 0
        op  = d.get("영업이익")
        net = d.get("당기순이익")
        dbt = d.get("부채총계") or 0
        eq  = d.get("자본총계") or 0
        return {
            "영업이익률": op  / rev * 100 if op  is not None and rev  else None,
            "순이익률":   net / rev * 100 if net is not None and rev  else None,
            "부채비율":   dbt / eq  * 100 if eq                       else None,
        }


# ── Company Selection Dialog ──────────────────────────────────────────────────
class SelectDialog(ctk.CTkToplevel):
    def __init__(self, parent, corps: list):
        super().__init__(parent)
        self.title("기업 선택")
        self.geometry("540x440")
        self.resizable(False, False)
        self.configure(fg_color=C["bg"])
        self.grab_set()
        self.result = None

        self.update_idletasks()
        px, py = parent.winfo_x(), parent.winfo_y()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        self.geometry(f"+{px + pw // 2 - 270}+{py + ph // 2 - 220}")

        # Header
        bar = ctk.CTkFrame(self, fg_color=C["blue"], corner_radius=0, height=56)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        ctk.CTkLabel(
            bar, text="분석할 기업을 선택하세요",
            font=ctk.CTkFont(FONT, 15, "bold"), text_color="white",
        ).place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(
            self,
            text=f"총 {len(corps)}개 기업이 검색되었습니다",
            font=ctk.CTkFont(FONT, 11),
            text_color=C["g1"],
        ).pack(pady=(10, 4))

        sf = ctk.CTkScrollableFrame(
            self, fg_color=C["bg"], corner_radius=0,
            scrollbar_button_color=C["g3"],
            scrollbar_button_hover_color=C["g2"],
        )
        sf.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        for corp in corps:
            row = ctk.CTkFrame(
                sf, fg_color=C["card"], corner_radius=10,
                border_width=1, border_color=C["g3"],
            )
            row.pack(fill="x", pady=3)
            row.columnconfigure(0, weight=1)

            info = ctk.CTkFrame(row, fg_color="transparent")
            info.grid(row=0, column=0, sticky="w", padx=14, pady=10)

            ctk.CTkLabel(
                info, text=corp["corp_name"],
                font=ctk.CTkFont(FONT, 13, "bold"),
                text_color=C["black"],
            ).pack(anchor="w")

            stock_txt = corp.get("stock_code") or "비상장"
            ctk.CTkLabel(
                info,
                text=f"종목코드: {stock_txt}  |  기업코드: {corp['corp_code']}",
                font=ctk.CTkFont(FONT, 11),
                text_color=C["g1"],
            ).pack(anchor="w")

            ctk.CTkButton(
                row, text="선택", width=64, height=32,
                corner_radius=8,
                fg_color=C["blue"], hover_color=C["bdark"],
                font=ctk.CTkFont(FONT, 12, "bold"),
                command=lambda c=corp: self._pick(c),
            ).grid(row=0, column=1, padx=(8, 14), pady=10)

    def _pick(self, corp):
        self.result = corp
        self.destroy()


# ── Section Card Widget ───────────────────────────────────────────────────────
class SectionCard(ctk.CTkFrame):
    """Displays a table: label | 2023 | 2024 (YoY) | 2025 (YoY)"""

    def __init__(self, parent, title: str):
        super().__init__(
            parent, fg_color=C["card"],
            corner_radius=16, border_width=1, border_color=C["g3"],
        )
        self._row_idx = 0
        self._build_header(title)

    def _build_header(self, title: str):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(16, 0))
        ctk.CTkLabel(
            hdr, text=title,
            font=ctk.CTkFont(FONT, 14, "bold"),
            text_color=C["black"], anchor="w",
        ).pack(side="left")

        ctk.CTkFrame(self, fg_color=C["g3"], height=1).pack(fill="x", padx=20, pady=(10, 0))

        # Column labels
        ch = ctk.CTkFrame(self, fg_color="transparent")
        ch.pack(fill="x", padx=20, pady=(6, 2))
        ch.columnconfigure((1, 2, 3), weight=1)

        ctk.CTkLabel(
            ch, text="항목",
            font=ctk.CTkFont(FONT, 11), text_color=C["g1"],
            width=110, anchor="w",
        ).grid(row=0, column=0, sticky="w")

        for i, yr in enumerate(YEARS):
            ctk.CTkLabel(
                ch, text=yr,
                font=ctk.CTkFont(FONT, 11), text_color=C["g1"],
                anchor="center",
            ).grid(row=0, column=i + 1, sticky="ew", padx=4)

    def add_row(self, label: str, vals: list, is_ratio: bool = False):
        bg = C["g4"] if self._row_idx % 2 == 0 else C["card"]

        rf = ctk.CTkFrame(self, fg_color=bg, corner_radius=8)
        rf.pack(fill="x", padx=12, pady=2)
        rf.columnconfigure((1, 2, 3), weight=1)

        ctk.CTkLabel(
            rf, text=label,
            font=ctk.CTkFont(FONT, 12), text_color=C["g1"],
            width=110, anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=(12, 4), pady=10)

        for i, v in enumerate(vals):
            cell = ctk.CTkFrame(rf, fg_color="transparent")
            cell.grid(row=0, column=i + 1, sticky="ew", padx=4, pady=8)

            val_str = fmt_pct(v) if is_ratio else fmt_num(v)
            ctk.CTkLabel(
                cell, text=val_str,
                font=ctk.CTkFont(FONT, 13, "bold"),
                text_color=C["black"], anchor="center",
            ).pack()

            if i > 0:
                txt, col = yoy(v, vals[i - 1])
                if txt:
                    ctk.CTkLabel(
                        cell, text=txt,
                        font=ctk.CTkFont(FONT, 10),
                        text_color=col or C["g1"],
                        anchor="center",
                    ).pack()

        self._row_idx += 1

    def finalize(self):
        ctk.CTkFrame(self, fg_color="transparent", height=10).pack()


# ── Main Application ──────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("기업 재무분석기")
        self.geometry("1100x800")
        self.minsize(880, 640)
        self.configure(fg_color=C["bg"])
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.dart    = DART()
        self.company = None
        self.results: dict = {}
        self._widgets: list = []

        self._build_ui()
        threading.Thread(
            target=self.dart.load, args=(self._on_load,), daemon=True,
        ).start()

    # ── UI Construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Navigation bar
        nav = ctk.CTkFrame(self, fg_color=C["card"], corner_radius=0, height=62)
        nav.pack(fill="x", side="top")
        nav.pack_propagate(False)

        logo = ctk.CTkFrame(nav, fg_color="transparent")
        logo.place(relx=0.5, rely=0.5, anchor="center")
        ctk.CTkFrame(
            logo, width=8, height=8, corner_radius=4, fg_color=C["blue"],
        ).pack(side="left", padx=(0, 8))
        ctk.CTkLabel(
            logo, text="기업 재무분석기",
            font=ctk.CTkFont(FONT, 18, "bold"), text_color=C["black"],
        ).pack(side="left")

        self.status_lbl = ctk.CTkLabel(
            nav, text="기업 코드 로딩 중...",
            font=ctk.CTkFont(FONT, 11), text_color=C["g1"],
        )
        self.status_lbl.place(relx=0.98, rely=0.5, anchor="e")

        # ── Search bar
        sb = ctk.CTkFrame(self, fg_color=C["card"], corner_radius=0, height=82)
        sb.pack(fill="x")
        sb.pack_propagate(False)

        si = ctk.CTkFrame(sb, fg_color="transparent")
        si.place(relx=0.5, rely=0.5, anchor="center")

        self.entry = ctk.CTkEntry(
            si,
            placeholder_text="기업명을 입력하세요  (예: 삼성전자, 카카오, LG화학)",
            width=440, height=48, corner_radius=12,
            border_color=C["g3"], fg_color=C["g4"],
            text_color=C["black"],
            font=ctk.CTkFont(FONT, 13),
        )
        self.entry.pack(side="left", padx=(0, 10))
        self.entry.bind("<Return>", lambda _: self._search())

        self.btn_search = ctk.CTkButton(
            si, text="분석 시작", width=116, height=48, corner_radius=12,
            fg_color=C["blue"], hover_color=C["bdark"],
            font=ctk.CTkFont(FONT, 13, "bold"),
            command=self._search,
        )
        self.btn_search.pack(side="left")

        self.btn_excel = ctk.CTkButton(
            si, text="⬇  Excel 저장", width=126, height=48, corner_radius=12,
            fg_color=C["green"], hover_color="#04A862",
            font=ctk.CTkFont(FONT, 13, "bold"),
            command=self._save_excel, state="disabled",
        )
        self.btn_excel.pack(side="left", padx=(10, 0))

        # Divider
        ctk.CTkFrame(self, fg_color=C["g3"], height=1).pack(fill="x")

        # ── Scrollable content area
        self.scroll = ctk.CTkScrollableFrame(
            self, fg_color=C["bg"], corner_radius=0,
            scrollbar_button_color=C["g3"],
            scrollbar_button_hover_color=C["g2"],
        )
        self.scroll.pack(fill="both", expand=True)

        self._show_placeholder()

    def _show_placeholder(self):
        ph = ctk.CTkFrame(self.scroll, fg_color="transparent")
        ph.pack(pady=100)
        self._widgets.append(ph)
        ctk.CTkLabel(ph, text="🏢", font=ctk.CTkFont(size=56)).pack()
        ctk.CTkLabel(
            ph, text="기업명을 입력하고 재무 분석을 시작하세요",
            font=ctk.CTkFont(FONT, 15), text_color=C["g2"],
        ).pack(pady=(14, 0))
        ctk.CTkLabel(
            ph, text="3개년(2023~2025) 재무제표 및 수익성 지표를 자동으로 분석합니다",
            font=ctk.CTkFont(FONT, 12), text_color=C["g3"],
        ).pack(pady=(6, 0))

    # ── Callbacks ─────────────────────────────────────────────────────────────
    def _on_load(self, ok: bool, msg: str = ""):
        def _update():
            if ok:
                self.status_lbl.configure(text="✓ 준비 완료", text_color=C["green"])
            else:
                self.status_lbl.configure(text=f"✗ 로드 실패", text_color=C["red"])
                self._toast(f"기업 코드 로드 실패: {msg}")
        self.after(0, _update)

    def _search(self):
        q = self.entry.get().strip()
        if not q:
            return
        if not self.dart.ready:
            self._toast("기업 코드를 로딩 중입니다. 잠시 후 다시 시도하세요.")
            return

        corps = self.dart.search(q)
        if not corps:
            self._toast(f"'{q}' 기업을 찾을 수 없습니다.")
            return

        exact = [c for c in corps if c["corp_name"] == q]
        pool  = exact if exact else corps

        if len(pool) == 1:
            self._run(pool[0])
        else:
            dlg = SelectDialog(self, pool[:30])
            self.wait_window(dlg)
            if dlg.result:
                self._run(dlg.result)

    def _run(self, corp: dict):
        self.company = corp
        self._clear()
        self._show_loading()
        threading.Thread(target=self._fetch, args=(corp,), daemon=True).start()

    def _fetch(self, corp: dict):
        res = {}
        for yr in YEARS:
            items, fs = self.dart.financials(corp["corp_code"], yr)
            if items:
                fin = self.dart.extract(items)
                rat = DART.ratios(fin)
                res[yr] = {**fin, **rat, "_fs": fs}
            else:
                res[yr] = None
        self.results = res
        self.after(0, lambda: self._display(corp["corp_name"], res))

    # ── Display ───────────────────────────────────────────────────────────────
    def _clear(self):
        for w in self._widgets:
            try:
                w.destroy()
            except Exception:
                pass
        self._widgets.clear()

    def _show_loading(self):
        f = ctk.CTkFrame(self.scroll, fg_color=C["card"], corner_radius=16)
        f.pack(fill="x", padx=40, pady=40)
        self._widgets.append(f)

        ctk.CTkLabel(
            f, text="DART 재무 데이터 조회 중...",
            font=ctk.CTkFont(FONT, 14), text_color=C["g1"],
        ).pack(pady=(36, 12))

        pb = ctk.CTkProgressBar(
            f, mode="indeterminate", width=300,
            fg_color=C["g3"], progress_color=C["blue"],
        )
        pb.pack(pady=(0, 12))
        pb.start()

        ctk.CTkLabel(
            f,
            text="2023 · 2024 · 2025 사업보고서를 순서대로 조회합니다",
            font=ctk.CTkFont(FONT, 11), text_color=C["g2"],
        ).pack(pady=(0, 36))

    def _display(self, corp_name: str, res: dict):
        self._clear()

        # ── Company banner
        banner = ctk.CTkFrame(
            self.scroll, fg_color=C["blue"], corner_radius=16,
        )
        banner.pack(fill="x", padx=40, pady=(24, 8))
        self._widgets.append(banner)

        bi = ctk.CTkFrame(banner, fg_color="transparent")
        bi.pack(fill="x", padx=24, pady=18)

        left_block = ctk.CTkFrame(bi, fg_color="transparent")
        left_block.pack(side="left")
        ctk.CTkLabel(
            left_block, text=corp_name,
            font=ctk.CTkFont(FONT, 22, "bold"),
            text_color="white", anchor="w",
        ).pack(anchor="w")

        # FS type label (연결/개별)
        fs_types = []
        for yr in YEARS:
            if res.get(yr):
                fs_types.append("연결" if res[yr].get("_fs") == "CFS" else "개별")
        if fs_types:
            fs_label = fs_types[0] + " 재무제표"
            ctk.CTkLabel(
                left_block,
                text=fs_label,
                font=ctk.CTkFont(FONT, 11),
                text_color="#FFFFFFAA",
                anchor="w",
            ).pack(anchor="w", pady=(2, 0))

        # Year availability dots
        dots = ctk.CTkFrame(bi, fg_color="transparent")
        dots.pack(side="right")
        for yr in YEARS:
            avail = res.get(yr) is not None
            dot_col = "white" if avail else "#FFFFFF44"
            ctk.CTkLabel(
                dots, text=f"● {yr}",
                font=ctk.CTkFont(FONT, 11), text_color=dot_col,
            ).pack(side="left", padx=6)

        # ── 손익계산서
        income = SectionCard(self.scroll, "손익계산서")
        income.pack(fill="x", padx=40, pady=8)
        self._widgets.append(income)
        for key in ("매출액", "영업이익", "당기순이익"):
            vals = [res[yr].get(key) if res.get(yr) else None for yr in YEARS]
            income.add_row(key, vals)
        income.finalize()

        # ── 재무상태표
        balance = SectionCard(self.scroll, "재무상태표")
        balance.pack(fill="x", padx=40, pady=8)
        self._widgets.append(balance)
        for key in ("자산총계", "부채총계", "자본총계"):
            vals = [res[yr].get(key) if res.get(yr) else None for yr in YEARS]
            balance.add_row(key, vals)
        balance.finalize()

        # ── 수익성 지표
        ratio = SectionCard(self.scroll, "수익성 지표")
        ratio.pack(fill="x", padx=40, pady=(8, 30))
        self._widgets.append(ratio)
        for key in ("영업이익률", "순이익률", "부채비율"):
            vals = [res[yr].get(key) if res.get(yr) else None for yr in YEARS]
            ratio.add_row(key, vals, is_ratio=True)
        ratio.finalize()

        self.btn_excel.configure(state="normal")

    # ── Excel Export ──────────────────────────────────────────────────────────
    def _save_excel(self):
        if not self.results or not self.company:
            return

        name = self.company["corp_name"]
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(os.path.expanduser("~/Desktop"), f"{name}_재무분석_{ts}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "재무분석"

        # ── Style definitions
        def mf(bold=False, color="191F28", size=11, italic=False):
            return Font(name="맑은 고딕", bold=bold, color=color, size=size, italic=italic)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        BORDER = Border(
            left=Side(style="thin", color="E5E8EB"),
            right=Side(style="thin", color="E5E8EB"),
            top=Side(style="thin", color="E5E8EB"),
            bottom=Side(style="thin", color="E5E8EB"),
        )
        CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT = Alignment(horizontal="left",   vertical="center")

        def cell(r, c, v="", font=None, bg=None, align=CTR):
            cc = ws.cell(r, c, v)
            if font:  cc.font      = font
            if bg:    cc.fill      = fill(bg)
            cc.alignment = align
            cc.border    = BORDER
            return cc

        # ── Title block
        ws.merge_cells("A1:G2")
        cell(1, 1, f"{name}  재무분석 보고서",
             mf(bold=True, color="FFFFFF", size=15), "3182F6", CTR)
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 20

        ws.merge_cells("A3:G3")
        cell(3, 1,
             f"분석일: {datetime.now().strftime('%Y년 %m월 %d일')}   |   출처: DART 전자공시시스템",
             mf(color="6B7684", size=10, italic=True), "FFFFFF", LEFT)
        ws.row_dimensions[3].height = 18

        r = 5
        SECTIONS = [
            ("손익계산서",  ["매출액", "영업이익", "당기순이익"],       False),
            ("재무상태표",  ["자산총계", "부채총계", "자본총계"],       False),
            ("수익성 지표", ["영업이익률", "순이익률", "부채비율"],     True),
        ]

        for sec_title, keys, is_ratio in SECTIONS:
            # Section title row
            ws.merge_cells(f"A{r}:G{r}")
            cell(r, 1, sec_title, mf(bold=True, color="FFFFFF", size=12), "4B9FFF", CTR)
            ws.row_dimensions[r].height = 28
            r += 1

            # Column header row
            for ci, h in enumerate(["항목", "2023", "2024", "YoY", "2025", "YoY", ""], 1):
                cell(r, ci, h, mf(bold=True, color="FFFFFF", size=10), "6B7684", CTR)
            ws.row_dimensions[r].height = 20
            r += 1

            for i, key in enumerate(keys):
                vals = [
                    self.results[yr].get(key) if self.results.get(yr) else None
                    for yr in YEARS
                ]
                bg = "F9FAFB" if i % 2 == 0 else "FFFFFF"
                ws.row_dimensions[r].height = 22

                cell(r, 1, key, mf(color="6B7684"), bg, LEFT)
                cell(r, 2, fmt_pct(vals[0]) if is_ratio else fmt_num(vals[0]),
                     mf(bold=True), bg, CTR)
                cell(r, 7, "", mf(), bg, CTR)

                for col_offset, idx in [(3, 1), (5, 2)]:
                    v = vals[idx]
                    cell(r, col_offset, fmt_pct(v) if is_ratio else fmt_num(v),
                         mf(bold=True), bg, CTR)
                    txt, color = yoy(v, vals[idx - 1])
                    if color == C["green"]:
                        yoy_font = mf(color="05C072", size=10)
                    elif color == C["red"]:
                        yoy_font = mf(color="F04452", size=10)
                    else:
                        yoy_font = mf(color="B0B8C1", size=10)
                    cell(r, col_offset + 1, txt or "-", yoy_font, bg, CTR)

                r += 1

            r += 1  # gap between sections

        # Column widths
        for col_letter, width in [
            ("A", 14), ("B", 14), ("C", 14), ("D", 11),
            ("E", 14), ("F", 11), ("G", 6),
        ]:
            ws.column_dimensions[col_letter].width = width

        wb.save(path)
        self._toast(f"저장 완료 → {os.path.basename(path)}")

    # ── Toast notification ────────────────────────────────────────────────────
    def _toast(self, msg: str):
        t = ctk.CTkToplevel(self)
        t.overrideredirect(True)
        t.attributes("-topmost", True)

        frame = ctk.CTkFrame(t, fg_color=C["black"], corner_radius=12)
        frame.pack()
        ctk.CTkLabel(
            frame, text=msg,
            font=ctk.CTkFont(FONT, 12), text_color="white",
        ).pack(padx=22, pady=13)

        self.update_idletasks()
        x = self.winfo_x() + self.winfo_width()  // 2 - 180
        y = self.winfo_y() + self.winfo_height() -  80
        t.geometry(f"+{x}+{y}")
        self.after(2800, t.destroy)


# ── Entry Point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
